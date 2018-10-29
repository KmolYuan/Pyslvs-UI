# -*- coding: utf-8 -*-

"""'dimensional_synthesis' module contains
dimensional synthesis functional interfaces.
"""

__author__ = "Yuan Chang"
__copyright__ = "Copyright (C) 2016-2018"
__license__ = "AGPL"
__email__ = "pyslvs@gmail.com"

from typing import (
    List,
    Tuple,
    Sequence,
    Dict,
    Callable,
    Union,
    Optional,
    Any,
)
from math import hypot
import csv
import pprint
from copy import deepcopy
from re import split as char_split
import openpyxl
from networkx import Graph
from core.QtModules import (
    pyqtSlot,
    QWidget,
    QModelIndex,
    QApplication,
    QMessageBox,
    QListWidgetItem,
    QIcon,
    QPixmap,
    QInputDialog,
    QDoubleSpinBox,
    QTableWidgetItem,
)
from core.graphics import PreviewCanvas, graph2vpoints
from core.io import str_before, str_between
from core.libs import expr_solving, VPoint
from core.synthesis import CollectionsDialog
from .ds_dialog import (
    GeneticPrams,
    FireflyPrams,
    defaultSettings,
    DifferentialPrams,
    AlgorithmType,
    AlgorithmOptionDialog,
    PathAdjustDialog,
    ProgressDialog,
    PreviewDialog,
    ChartDialog,
)
from core import main_window as mw
from .Ui_dimension_widget import Ui_Form

__all__ = ['DimensionalSynthesis']


class DimensionalSynthesis(QWidget, Ui_Form):

    """Dimensional synthesis widget.

    User can run the dimensional synthesis here.
    """

    def __init__(self, parent: 'mw.MainWindow'):
        """Reference names:

        + Iteration collections.
        + Result data.
        + Main window function references.
        """
        super(DimensionalSynthesis, self).__init__(parent)
        self.setupUi(self)

        self.mech_params = {}
        self.path = {}

        # Some reference of 'collections'.
        self.collections = parent.CollectionTabPage.TriangularIterationWidget.collections
        self.getCollection = parent.getCollection
        self.inputFrom = parent.inputFrom
        self.unsaveFunc = parent.workbookNoSave
        self.mergeResult = parent.mergeResult
        self.updateRanges = parent.MainCanvas.updateRanges
        self.setSolvingPath = parent.MainCanvas.setSolvingPath

        # Data and functions.
        self.__mechanism_data = []
        self.alg_options = {}
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.__set_algorithm_default()

        def get_solutions_func() -> str:
            """For preview canvas."""
            if 'Expression' in self.mech_params:
                return self.mech_params['Expression']
            else:
                return ""

        self.PreviewCanvas = PreviewCanvas(get_solutions_func, self)
        self.preview_layout.addWidget(self.PreviewCanvas)
        self.show_solutions.clicked.connect(self.PreviewCanvas.setShowSolutions)

        # Splitter
        self.up_splitter.setSizes([80, 100])
        self.down_splitter.setSizes([20, 80])

        # Table widget column width.
        self.parameter_list.setColumnWidth(0, 75)
        self.parameter_list.setColumnWidth(1, 75)
        self.parameter_list.setColumnWidth(2, 70)
        self.parameter_list.setColumnWidth(3, 70)
        self.parameter_list.setColumnWidth(4, 80)

        self.clear()

    def clear(self):
        """Clear all sub-widgets."""
        self.__mechanism_data.clear()
        self.result_list.clear()
        self.__clear_settings()
        self.__has_result()

    def __clear_settings(self):
        """Clear sub-widgets that contain the setting."""
        self.__clear_path(ask=False)
        self.path.clear()
        self.mech_params.clear()
        self.PreviewCanvas.clear()
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.profile_name.setText("No setting")
        self.type2.setChecked(True)
        self.parameter_list.setRowCount(0)
        self.target_points.clear()
        self.Expression.clear()
        self.Link_expr.clear()
        self.updateRange()
        self.__able_to_generate()

    def mechanism_data(
        self,
        index: Optional[int] = None
    ) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
        """Return the index of mechanism data."""
        if index is not None:
            return self.__mechanism_data[index]
        return self.__mechanism_data

    @pyqtSlot(name='on_clear_button_clicked')
    def __user_clear(self):
        if self.profile_name.text() == "No setting":
            return
        reply = QMessageBox.question(
            self,
            "Clear setting",
            "Do you want to clear the setting?"
        )
        if reply == QMessageBox.Yes:
            self.__clear_settings()

    def loadResults(self, mechanism_data: Sequence[Dict[str, Any]]):
        """Append results of workbook database to memory."""
        for e in mechanism_data:
            self.__mechanism_data.append(e)
            self.__add_result(e)

    def __current_path_changed(self):
        """Call the canvas to update to current target path."""
        self.setSolvingPath({
            name: tuple(path) for name, path in self.path.items()
        })
        self.__able_to_generate()

    def currentPath(self) -> List[Tuple[float, float]]:
        """Return the pointer of current target path."""
        item = self.target_points.currentItem()
        if item:
            return self.path[item.text()]
        else:
            return []

    @pyqtSlot(str, name='on_target_points_currentTextChanged')
    def __set_target(self, _: str):
        """Switch to the current target path."""
        self.path_list.clear()
        for x, y in self.currentPath():
            self.path_list.addItem(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @pyqtSlot(name='on_path_clear_clicked')
    def __clear_path(self, *, ask: bool = True):
        """Clear the current target path."""
        if ask:
            reply = QMessageBox.question(
                self,
                "Clear path",
                "Are you sure to clear the current path?"
            )
            if reply != QMessageBox.Yes:
                return
        self.currentPath().clear()
        self.path_list.clear()
        self.__current_path_changed()

    @pyqtSlot(name='on_path_copy_clicked')
    def __copy_path(self):
        """Copy the current path coordinates to clipboard."""
        QApplication.clipboard().setText('\n'.join(
            f"{x},{y}" for x, y in self.currentPath()
        ))

    @pyqtSlot(name='on_path_paste_clicked')
    def __paste_path(self):
        """Paste path data from clipboard."""
        self.__read_path_from_csv(char_split("[;,\n]", QApplication.clipboard().text()))

    @pyqtSlot(name='on_import_csv_button_clicked')
    def __import_csv(self):
        """Paste path data from a text file."""
        file_name = self.inputFrom(
            "Path data",
            ["Text file (*.txt)", "Comma-Separated Values (*.csv)"]
        )
        if not file_name:
            return
        data = []
        with open(file_name, newline='') as stream:
            for row in csv.reader(stream, delimiter=' ', quotechar='|'):
                data += " ".join(row).split(',')
        self.__read_path_from_csv(data)

    def __read_path_from_csv(self, raw_data: List[str]):
        """Turn string to float then add them to current target path."""
        try:
            data = [
                (round(float(raw_data[i]), 4), round(float(raw_data[i + 1]), 4))
                for i in range(0, len(raw_data), 2)
            ]
        except (IndexError, ValueError):
            QMessageBox.warning(
                self,
                "File error",
                "Wrong format.\nIt should be look like this:"
                "\n0.0,0.0[\\n]" * 3
            )
        else:
            for x, y in data:
                self.addPoint(x, y)

    @pyqtSlot(name='on_import_xlsx_button_clicked')
    def __import_xlsx(self):
        """Paste path data from a Excel file."""
        file_name = self.inputFrom(
            "Excel file",
            ["Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)"]
        )
        if not file_name:
            return
        wb = openpyxl.load_workbook(file_name)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        data = []
        # Keep finding until there is no value.
        i = 1
        while True:
            x = ws.cell(row=i, column=1).value
            y = ws.cell(row=i, column=2).value
            if None in {x, y}:
                break
            try:
                data.append((round(float(x), 4), round(float(y), 4)))
            except (IndexError, AttributeError):
                QMessageBox.warning(
                    self,
                    "File error",
                    "Wrong format.\n"
                    "The data sheet seems to including non-digital cell."
                )
                break
            i += 1
        for x, y in data:
            self.addPoint(x, y)

    @pyqtSlot(name='on_path_adjust_button_clicked')
    def __adjust_path(self):
        """Show up path adjust dialog and
        get back the changes of current target path.
        """
        dlg = PathAdjustDialog(self)
        dlg.show()
        if not dlg.exec_():
            return
        self.__clear_path(ask=False)
        for e in dlg.r_path:
            self.addPoint(e[0], e[1])
        self.__current_path_changed()

    def addPoint(self, x: float, y: float):
        """Add path data to list widget and current target path."""
        x = round(x, 4)
        y = round(y, 4)
        self.currentPath().append((x, y))
        self.path_list.addItem(f"({x:.04f}, {y:.04f})")
        self.path_list.setCurrentRow(self.path_list.count() - 1)
        self.__current_path_changed()

    @pyqtSlot(float, float)
    def setPoint(self, x: float, y: float):
        """Set the coordinate of current target path."""
        if not self.edit_target_point_button.isChecked():
            return
        for i, (cx, cy) in enumerate(self.currentPath()):
            if hypot(x - cx, y - cy) < 3:
                index = i
                self.path_list.setCurrentRow(index)
                break
        else:
            return
        self.currentPath()[index] = (x, y)
        self.path_list.item(index).setText(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @pyqtSlot(name='on_close_path_clicked')
    def __close_path(self):
        """Add a the last point same as first point."""
        current_path = self.currentPath()
        if (self.path_list.count() > 1) and (current_path[0] != current_path[-1]):
            self.addPoint(*current_path[0])

    @pyqtSlot(name='on_point_up_clicked')
    def __move_up_point(self):
        """Target point move up."""
        row = self.path_list.currentRow()
        if not ((row > 0) and (self.path_list.count() > 1)):
            return
        path = self.currentPath()
        path.insert(row - 1, (path[row][0], path[row][1]))
        del path[row + 1]
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row - 1, f"({x}, {y})")
        self.path_list.takeItem(row + 1)
        self.path_list.setCurrentRow(row - 1)
        self.__current_path_changed()

    @pyqtSlot(name='on_point_down_clicked')
    def __move_down_point(self):
        """Target point move down."""
        row = self.path_list.currentRow()
        if not (
            (row < self.path_list.count() - 1) and
            (self.path_list.count() > 1)
        ):
            return
        path = self.currentPath()
        path.insert(row + 2, (path[row][0], path[row][1]))
        del path[row]
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row + 2, f"({x}, {y})")
        self.path_list.takeItem(row)
        self.path_list.setCurrentRow(row + 1)
        self.__current_path_changed()

    @pyqtSlot(name='on_point_delete_clicked')
    def __delete_point(self):
        """Delete a target point."""
        row = self.path_list.currentRow()
        if not row > -1:
            return
        del self.currentPath()[row]
        self.path_list.takeItem(row)
        self.__current_path_changed()

    def __able_to_generate(self):
        """Set button enable if all the data are already."""
        self.pointNum.setText(
            "<p><span style=\"font-size:12pt;"
            f"color:#00aa00;\">{self.path_list.count()}</span></p>"
        )
        n = (
            bool(self.mech_params) and
            (self.path_list.count() > 1) and
            bool(self.Expression.text())
        )
        self.path_adjust_button.setEnabled(n)
        self.synthesis_button.setEnabled(n)

    @pyqtSlot(name='on_synthesis_button_clicked')
    def __synthesis(self):
        """Start synthesis."""
        # Check if the number of target points are same.
        length = -1
        for path in self.path.values():
            if length < 0:
                length = len(path)
            if len(path) != length:
                QMessageBox.warning(
                    self,
                    "Target Error",
                    "The length of target paths should be the same."
                )
                return
        # Get the algorithm type.
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        else:
            type_num = AlgorithmType.DE
        # Deep copy it so the pointer will not the same.
        mech_params = deepcopy(self.mech_params)
        mech_params['Target'] = deepcopy(self.path)

        def name_in_table(target_name: str) -> int:
            """Find a target_name and return the row from the table."""
            for r in range(self.parameter_list.rowCount()):
                if self.parameter_list.item(r, 0).text() == target_name:
                    return r
            return -1

        for key in ('Driver', 'Follower'):
            for name in mech_params[key]:
                row = name_in_table(name)
                mech_params[key][name] = (
                    self.parameter_list.cellWidget(row, 2).value(),
                    self.parameter_list.cellWidget(row, 3).value(),
                    self.parameter_list.cellWidget(row, 4).value(),
                )

        # Start progress dialog.
        dlg = ProgressDialog(type_num, mech_params, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            return
        for m in dlg.mechanisms:
            self.__mechanism_data.append(m)
            self.__add_result(m)
        self.__set_time(dlg.time_spend)
        self.unsaveFunc()
        QMessageBox.information(
            self,
            "Dimensional Synthesis",
            "Your tasks is all completed."
        )
        print("Finished.")

    def __set_time(self, time: float):
        """Set the time label."""
        self.timeShow.setText(
            "<html><head/><body><p><span style=\"font-size:16pt\">"
            f"{time // 60}[min] {time % 60:.02f}[s]"
            "</span></p></body></html>"
        )

    def __add_result(self, result: Dict[str, Any]):
        """Add result items, except add to the list."""
        item = QListWidgetItem(result['Algorithm'])
        interrupt = result['interrupted']
        if interrupt == 'False':
            interrupt_icon = "task_completed.png"
        elif interrupt == 'N/A':
            interrupt_icon = "question.png"
        else:
            interrupt_icon = "interrupted.png"
        item.setIcon(QIcon(QPixmap(f":/icons/{interrupt_icon}")))
        if interrupt == 'False':
            interrupt_text = "No interrupt."
        else:
            interrupt_text = f"Interrupt at: {interrupt}"
        text = f"{result['Algorithm']} ({interrupt_text})"
        if interrupt == 'N/A':
            text += "\n※Completeness is unknown."
        item.setToolTip(text)
        self.result_list.addItem(item)

    @pyqtSlot(name='on_delete_button_clicked')
    def __delete_result(self):
        """Delete a result."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        reply = QMessageBox.question(
            self,
            "Delete",
            "Delete this result from list?"
        )
        if reply != QMessageBox.Yes:
            return
        del self.__mechanism_data[row]
        self.result_list.takeItem(row)
        self.unsaveFunc()
        self.__has_result()

    @pyqtSlot(QModelIndex, name='on_result_list_clicked')
    def __has_result(self, *_: QModelIndex):
        """Set enable if there has any result."""
        enable = self.result_list.currentRow() > -1
        for button in (
            self.merge_button,
            self.delete_button,
            self.result_load_settings,
            self.result_chart,
            self.result_clipboard
        ):
            button.setEnabled(enable)

    @pyqtSlot(QModelIndex, name='on_result_list_doubleClicked')
    def __show_result(self, _: QModelIndex):
        """Double click result item can show up preview dialog."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        dlg = PreviewDialog(self.__mechanism_data[row], self.__get_path(row), self)
        dlg.show()
        dlg.exec_()

    @pyqtSlot(name='on_merge_button_clicked')
    def __merge_result(self):
        """Merge mechanism into main canvas."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        reply = QMessageBox.question(
            self,
            "Merge",
            "Merge this result to your canvas?"
        )
        if reply == QMessageBox.Yes:
            self.mergeResult(row, self.__get_path(row))

    def __get_path(self, row: int) -> List[List[Tuple[float, float]]]:
        """Using result data to generate paths of mechanism."""
        result = self.__mechanism_data[row]

        exprs = []
        for expr in result['Expression'].split(';'):
            func = str_before(expr, '[')
            params = str_between(expr, '[', ']').split(',')
            target = str_between(expr, '(', ')')
            params.insert(0, func)
            params.append(target)
            exprs.append(tuple(params))

        pos = {}
        mapping = {}
        p_count = 0
        for n in result['pos']:
            try:
                pos[n] = result[f'P{n}']
            except KeyError:
                pos[n] = result['pos'][n]
            if n not in result['same']:
                mapping[p_count] = f'P{n}'
                p_count += 1

        vpoints = graph2vpoints(
            Graph(result['Graph']),
            pos,
            result['cus'],
            result['same']
        )
        vpoint_count = len(vpoints)

        path = []
        for i in range(vpoint_count):
            path.append([])

        # Cumulative angle
        i_count = sum(1 for e in exprs if e[0] == 'PLAP')
        angles_cum = [0.] * i_count
        nan = float('nan')
        for interval in (3, -3):
            # Driver pointer
            dp = 0
            angles = [0.] * i_count
            while dp < i_count:
                try:
                    solved_result = expr_solving(exprs, mapping, vpoints, angles)
                except RuntimeError:
                    # Update with error sign.
                    for i in range(vpoint_count):
                        path[i].append((nan, nan))
                    # Back to last feasible solution.
                    angles[dp] -= interval
                    dp += 1
                else:
                    # Update with result.
                    for i in range(vpoint_count):
                        if vpoints[i].type in {VPoint.P, VPoint.RP}:
                            slot: Tuple[float, float] = solved_result[i][0]
                            pin: Tuple[float, float] = solved_result[i][1]
                            path[i].append(solved_result[i][1])
                            vpoints[i].move(slot, pin)
                        else:
                            path[i].append(solved_result[i])
                            vpoints[i].move(solved_result[i])
                    angles[dp] += interval
                    angles[dp] %= 360
                    angles_cum[dp] += abs(interval)
                    if angles_cum[dp] > 360:
                        angles[dp] -= interval
                        dp += 1
        return path

    @pyqtSlot(name='on_result_chart_clicked')
    def __show_result_chart(self):
        """Show up the chart dialog."""
        dlg = ChartDialog("Convergence Value", self.__mechanism_data, self)
        dlg.show()
        dlg.exec_()

    @pyqtSlot(name='on_result_clipboard_clicked')
    def __copy_result_text(self):
        """Copy pretty print result as text."""
        QApplication.clipboard().setText(
            pprint.pformat(self.__mechanism_data[self.result_list.currentRow()])
        )

    @pyqtSlot(name='on_save_profile_clicked')
    def __save_profile(self):
        """Save as new profile to collection widget."""
        if not self.mech_params:
            return
        name, ok = QInputDialog.getText(
            self,
            "Profile name",
            "Please enter the profile name:"
        )
        if not ok:
            return
        i = 0
        while (not name) and (name not in self.collections):
            name = f"Structure_{i}"
            i += 1

        mech_params = deepcopy(self.mech_params)
        for key in ['Driver', 'Follower', 'Target']:
            for mp in mech_params[key]:
                mech_params[key][mp] = None

        self.collections[name] = mech_params
        self.unsaveFunc()

    @pyqtSlot(name='on_load_profile_clicked')
    def __load_profile(self):
        """Load profile from collections dialog."""
        dlg = CollectionsDialog(
            self.collections,
            self.getCollection,
            self
        )
        dlg.show()
        if dlg.exec_():
            self.__set_profile(dlg.name(), dlg.params())

    def __set_profile(self, profile_name: str, params: Dict[str, Any]):
        """Set profile to sub-widgets."""
        self.__clear_settings()
        self.profile_name.setText(profile_name)
        self.mech_params = deepcopy(params)
        self.Expression.setText(self.mech_params['Expression'])
        self.Link_expr.setText(self.mech_params['Link_expr'])
        for name in sorted(self.mech_params['Target']):
            self.target_points.addItem(name)
            path = self.mech_params['Target'][name]
            if path:
                self.path[name] = path.copy()
            else:
                self.path[name] = []
        if self.target_points.count():
            self.target_points.setCurrentRow(0)

        gj = {}
        for key in ('Driver', 'Follower'):
            gj.update(self.mech_params[key])

        link_list = set()
        angle_list = set()
        for expr in self.mech_params['Expression'].split(';'):
            for e in str_between(expr, '[', ']').split(','):
                if e.startswith('L'):
                    link_list.add(e)
                if e.startswith('a'):
                    angle_list.add(e)
        link_count = len(link_list)
        angle_count = len(angle_list)

        self.parameter_list.setRowCount(0)
        self.parameter_list.setRowCount(len(gj) + link_count + angle_count)

        def spinbox(
            v: float,
            *,
            minimum: float = 0,
            maximum: float = 1000000.0,
            prefix: bool = False
        ) -> QDoubleSpinBox:
            double_spinbox = QDoubleSpinBox()
            double_spinbox.setMinimum(minimum)
            double_spinbox.setMaximum(maximum)
            double_spinbox.setSingleStep(10.0)
            double_spinbox.setValue(v)
            if prefix:
                double_spinbox.setPrefix("±")
            return double_spinbox

        row = 0
        for name in sorted(gj):
            coord = gj[name]
            self.parameter_list.setItem(row, 0, QTableWidgetItem(name))
            if name in self.mech_params['Driver']:
                role = 'Driver'
            else:
                role = 'Follower'
            self.parameter_list.setItem(row, 1, QTableWidgetItem(role))
            x, y = self.mech_params['pos'][int(name.replace('P', ''))]
            s1 = spinbox(coord[0] if coord else x, minimum=-1000000.0)
            s2 = spinbox(coord[1] if coord else y, minimum=-1000000.0)
            s3 = spinbox(coord[2] if coord else 50., prefix=True)
            self.parameter_list.setCellWidget(row, 2, s1)
            self.parameter_list.setCellWidget(row, 3, s2)
            self.parameter_list.setCellWidget(row, 4, s3)
            # Signal connections.
            for s in (s1, s2, s3):
                s.valueChanged.connect(self.updateRange)
            row += 1

        def set_by_center(
            index: int,
            get_range: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by center."""

            @pyqtSlot(float)
            def func(value: float):
                half_range = get_range() / 2
                self.mech_params['upper'][index] = value + half_range
                self.mech_params['lower'][index] = value - half_range

            return func

        def set_by_range(
            index: int,
            get_value: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by range."""

            @pyqtSlot(float)
            def func(value: float):
                center = get_value()
                half_range = value / 2
                self.mech_params['upper'][index] = center + half_range
                self.mech_params['lower'][index] = center - half_range

            return func

        for name in ('upper', 'lower'):
            if name not in self.mech_params:
                self.mech_params[name] = [0.] * (link_count + angle_count)

        for i, name in enumerate(sorted(link_list) + sorted(angle_list)):
            name_item = QTableWidgetItem(name)
            name_item.setToolTip(name)
            self.parameter_list.setItem(row, 0, name_item)
            self.parameter_list.setItem(row, 1, QTableWidgetItem('Limit'))
            # Set values (it will be same if not in the 'mech_params').
            upper = self.mech_params['upper'][i]
            if upper == 0:
                upper = 100. if name in link_list else 360.
            lower = self.mech_params['lower'][i]
            if lower == 0 and name in link_list:
                lower = 5.
            self.mech_params['upper'][i] = upper
            self.mech_params['lower'][i] = lower
            # Spin box.
            error_range = upper - lower
            default_value = error_range / 2 + lower
            if name in link_list:
                s1 = spinbox(default_value)
            else:
                s1 = spinbox(default_value, maximum=360.)
            self.parameter_list.setCellWidget(row, 2, s1)
            s2 = spinbox(error_range, prefix=True)
            self.parameter_list.setCellWidget(row, 4, s2)
            # Signal connections.
            s1.valueChanged.connect(set_by_center(i, s2.value))
            s2.valueChanged.connect(set_by_range(i, s1.value))
            row += 1

        self.PreviewCanvas.from_profile(self.mech_params)
        self.updateRange()
        self.alg_options.update(self.mech_params['settings'])
        self.__able_to_generate()
        if not self.Expression.text():
            QMessageBox.warning(
                self,
                "Profile cannot use",
                "This profile has no any solutions, "
                "you can set it in the \"Triangular iteration\" page."
            )

    @pyqtSlot(name='on_result_load_settings_clicked')
    def __load_result_settings(self):
        """Load settings from a result."""
        self.__has_result()
        row = self.result_list.currentRow()
        if not row > -1:
            return
        self.__clear_settings()
        result = self.__mechanism_data[row]
        if result['Algorithm'] == str(AlgorithmType.RGA):
            self.type0.setChecked(True)
        elif result['Algorithm'] == str(AlgorithmType.Firefly):
            self.type1.setChecked(True)
        elif result['Algorithm'] == str(AlgorithmType.DE):
            self.type2.setChecked(True)
        # Copy to mechanism params.
        self.__set_profile("External setting", result)
        self.__set_time(result['time'])
        # Load settings.
        self.alg_options.clear()
        self.alg_options.update(result['settings'])

    @pyqtSlot(name='on_type0_clicked')
    @pyqtSlot(name='on_type1_clicked')
    @pyqtSlot(name='on_type2_clicked')
    def __set_algorithm_default(self):
        """Set the algorithm settings to default."""
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        if self.type0.isChecked():
            self.alg_options.update(GeneticPrams)
        elif self.type1.isChecked():
            self.alg_options.update(FireflyPrams)
        elif self.type2.isChecked():
            self.alg_options.update(DifferentialPrams)

    @pyqtSlot(name='on_advance_button_clicked')
    def __show_advance(self):
        """Get the settings from advance dialog."""
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        else:
            type_num = AlgorithmType.DE
        dlg = AlgorithmOptionDialog(type_num, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            return
        self.alg_options['report'] = dlg.report.value()
        if dlg.maxGen_option.isChecked():
            self.alg_options['maxGen'] = dlg.maxGen.value()
        elif dlg.minFit_option.isChecked():
            self.alg_options['minFit'] = dlg.minFit.value()
        elif dlg.maxTime_option.isChecked():
            # Three spinbox value translate to second.
            self.alg_options['maxTime'] = (
                dlg.maxTime_h.value() * 3600 +
                dlg.maxTime_m.value() * 60 +
                dlg.maxTime_s.value()
            )

        def from_table(row: int) -> Union[int, float]:
            """Get algorithm data from table."""
            return dlg.alg_table.cellWidget(row, 1).value()

        pop_size = dlg.pop_size.value()
        if type_num == AlgorithmType.RGA:
            self.alg_options['nPop'] = pop_size
            for i, tag in enumerate(('pCross', 'pMute', 'pWin', 'bDelta')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.Firefly:
            self.alg_options['n'] = pop_size
            for i, tag in enumerate(('alpha', 'betaMin', 'gamma', 'beta0')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.DE:
            self.alg_options['NP'] = pop_size
            for i, tag in enumerate(('strategy', 'F', 'CR')):
                self.alg_options[tag] = from_table(i)

    @pyqtSlot()
    def updateRange(self):
        """Update range values to main canvas."""

        def t(x: int, y: int):
            item = self.parameter_list.item(x, y)
            if item:
                return item.text()
            else:
                return self.parameter_list.cellWidget(x, y).value()

        self.updateRanges({
            t(row, 0): (t(row, 2), t(row, 3), t(row, 4))
            for row in range(self.parameter_list.rowCount())
            if t(row, 1) in {'Follower', 'Driver'}
        })

    @pyqtSlot(name='on_expr_copy_clicked')
    def __copy_expr(self):
        """Copy profile expression."""
        text = self.Expression.text()
        if text:
            QApplication.clipboard().setText(text)

    @pyqtSlot(name='on_link_expr_copy_clicked')
    def __copy_link_expr(self):
        """Copy profile link expression."""
        text = self.Link_expr.text()
        if text:
            QApplication.clipboard().setText(text)
