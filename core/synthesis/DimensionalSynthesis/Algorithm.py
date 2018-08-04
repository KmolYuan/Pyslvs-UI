# -*- coding: utf-8 -*-

"""The widget of 'Dimensional synthesis' tab."""

__author__ = "Yuan Chang"
__copyright__ = "Copyright (C) 2016-2018"
__license__ = "AGPL"
__email__ = "pyslvs@gmail.com"

from typing import (
    List,
    Dict,
    Tuple,
    Callable,
    Any,
    Union,
    Optional,
)
import csv
import pprint
from copy import deepcopy
from re import split as charSplit
import openpyxl
from networkx import Graph
from core.QtModules import (
    pyqtSlot,
    QWidget,
    QApplication,
    QMessageBox,
    QListWidgetItem,
    QIcon,
    QPixmap,
    QModelIndex,
    QInputDialog,
    QDoubleSpinBox,
    QTableWidgetItem,
)
from core.graphics import (
    PreviewCanvas,
    graph2vpoints,
)
from core.io import (
    strbetween,
    strbefore,
)
from core.libs import expr_solving
from core.synthesis import CollectionsDialog
from .DimensionalSynthesis_dialog import (
    GeneticPrams,
    FireflyPrams,
    defaultSettings,
    DifferentialPrams,
    AlgorithmType,
    AlgorithmOptionDialog,
    PathAdjustDialog,
    ProgressDialog,
    PreviewDialog,
    ChartDialog,
)
from .Ui_Algorithm import Ui_Form


class DimensionalSynthesis(QWidget, Ui_Form):
    
    """Dimensional synthesis widget.
    
    User can run the dimensional synthesis here.
    """
    
    def __init__(self, parent: QWidget):
        """Reference names:
        
        + Iteration collections.
        + Result data.
        + Main window function references.
        """
        super(DimensionalSynthesis, self).__init__(parent)
        self.setupUi(self)
        
        self.mech_params = {}
        self.path = {}
        
        #Some reference of 'collections'.
        self.collections = parent.CollectionTabPage.TriangularIterationWidget.collections
        self.getCollection = parent.getCollection
        self.inputFrom = parent.inputFrom
        self.unsaveFunc = parent.workbookNoSave
        self.mergeResult = parent.mergeResult
        self.updateRanges = parent.MainCanvas.updateRanges
        self.setSolvingPath = parent.MainCanvas.setSolvingPath
        
        #Data and functions.
        self.__mechanism_data = []
        self.alg_options = {}
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.__setAlgorithmToDefault()
        
        def get_solutions_func() -> Tuple[str]:
            """For preview canvas."""
            try:
                return self.mech_params['Expression']
            except KeyError:
                return ()
        
        self.PreviewCanvas = PreviewCanvas(get_solutions_func, self)
        self.preview_layout.addWidget(self.PreviewCanvas)
        self.show_solutions.clicked.connect(self.PreviewCanvas.setShowSolutions)
        
        #Splitter
        self.up_splitter.setSizes([80, 100])
        self.down_splitter.setSizes([20, 80])
        
        #Table widget column width.
        self.parameter_list.setColumnWidth(0, 75)
        self.parameter_list.setColumnWidth(1, 75)
        self.parameter_list.setColumnWidth(2, 70)
        self.parameter_list.setColumnWidth(3, 70)
        self.parameter_list.setColumnWidth(4, 80)
        
        self.clear()
    
    def clear(self):
        """Clear all sub-widgets."""
        self.__mechanism_data.clear()
        self.result_list.clear()
        self.__clearSettings()
        self.__hasResult()
    
    def __clearSettings(self):
        """Clear sub-widgets that contain the setting."""
        self.__clearPath(ask=False)
        self.path.clear()
        self.mech_params.clear()
        self.PreviewCanvas.clear()
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.profile_name.setText("No setting")
        self.type2.setChecked(True)
        self.parameter_list.setRowCount(0)
        self.target_points.clear()
        self.Expression.clear()
        self.Link_expr.clear()
        self.updateRange()
        self.__ableToGenerate()
    
    def mechanismData(self,
        index: Optional[int] = None
    ) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
        """Return the index of mechanism data."""
        if index is not None:
            return self.__mechanism_data[index]
        return self.__mechanism_data
    
    def on_clear_button_clicked(self):
        if self.profile_name.text() == "No setting":
            return
        reply = QMessageBox.question(
            self,
            "Clear setting",
            "Do you want to clear the setting?"
        )
        if reply == QMessageBox.Yes:
            self.__clearSettings()
    
    def loadResults(self,
        __mechanism_data: List[Dict[str, Any]]
    ):
        """Append results of workbook database to memory."""
        for e in __mechanism_data:
            self.__mechanism_data.append(e)
            self.__addResult(e)
    
    def __currentPathChanged(self):
        """Call the canvas to update to current target path."""
        self.setSolvingPath({
            name: tuple(path) for name, path in self.path.items()
        })
        self.__ableToGenerate()
    
    def currentPath(self) -> List[Tuple[float, float]]:
        """Return the pointer of current target path."""
        item = self.target_points.currentItem()
        if item:
            return self.path[item.text()]
        else:
            return []
    
    @pyqtSlot(str, name='on_target_points_currentTextChanged')
    def __setTarget(self, text: Optional[str] = None):
        """Switch to the current target path."""
        self.path_list.clear()
        for x, y in self.currentPath():
            self.path_list.addItem("({:.04f}, {:.04f})".format(x, y))
        self.__currentPathChanged()
    
    @pyqtSlot(name='on_path_clear_clicked')
    def __clearPath(self, *, ask: bool = True):
        """Clear the current target path."""
        if ask:
            reply = QMessageBox.question(self,
                "Clear path",
                "Are you sure to clear the current path?"
            )
            if reply != QMessageBox.Yes:
                return
        self.currentPath().clear()
        self.path_list.clear()
        self.__currentPathChanged()
    
    @pyqtSlot(name='on_path_copy_clicked')
    def __copyPath(self):
        """Copy the current path coordinates to clipboard."""
        QApplication.clipboard().setText('\n'.join(
            "{},{}".format(x, y)
            for x, y in self.currentPath()
        ))
    
    @pyqtSlot(name='on_path_paste_clicked')
    def __pastePath(self):
        """Paste path data from clipboard."""
        self.__readPathFromCSV(charSplit(";|,|\n", QApplication.clipboard().text()))
    
    @pyqtSlot(name='on_import_csv_button_clicked')
    def __importCSV(self):
        """Paste path data from a text file."""
        file_name = self.inputFrom(
            "Path data",
            ["Text file (*.txt)",
            "Comma-Separated Values (*.csv)"]
        )
        if not file_name:
            return
        data = []
        with open(file_name, newline='') as stream:
            reader = csv.reader(stream, delimiter=' ', quotechar='|')
            for row in reader:
                data += ' '.join(row).split(',')
        self.__readPathFromCSV(data)
    
    def __readPathFromCSV(self, data: List[str]):
        """Trun STR to FLOAT then add them to current target path."""
        try:
            data = [
                (round(float(data[i]), 4), round(float(data[i + 1]), 4))
                for i in range(0, len(data), 2)
            ]
        except:
            QMessageBox.warning(self, "File error",
                "Wrong format.\nIt should be look like this:" +
                "\n0.0,0.0[\\n]" * 3
            )
        else:
            for e in data:
                self.addPoint(e[0], e[1])
    
    @pyqtSlot(name='on_import_xlsx_button_clicked')
    def __importXLSX(self):
        """Paste path data from a Excel file."""
        file_name = self.inputFrom(
            "Excel file",
            ["Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)"]
        )
        if not file_name:
            return
        wb = openpyxl.load_workbook(file_name)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        data = []
        #Keep finding until there is no value.
        i = 1
        while True:
            x = ws.cell(row=i, column=1).value
            y = ws.cell(row=i, column=2).value
            if x == None or y == None:
                break
            try:
                data.append((round(float(x), 4), round(float(y), 4)))
            except:
                QMessageBox.warning(self,
                    "File error",
                    "Wrong format.\n" +
                    "The datasheet seems to including non-digital cell."
                )
                break
            i += 1
        for x, y in data:
            self.addPoint(x, y)
    
    @pyqtSlot(name='on_path_adjust_button_clicked')
    def __adjustPath(self):
        """Show up path adjust dialog and
        get back the changes of current target path.
        """
        dlg = PathAdjustDialog(self)
        dlg.show()
        if not dlg.exec_():
            return
        self.__clearPath(ask=False)
        for e in dlg.r_path:
            self.addPoint(e[0], e[1])
        self.__currentPathChanged()
    
    def addPoint(self, x: float, y: float):
        """Add path data to list widget and
        current target path.
        """
        x = round(x, 4)
        y = round(y, 4)
        self.currentPath().append((x, y))
        self.path_list.addItem("({:.04f}, {:.04f})".format(x, y))
        self.__currentPathChanged()
    
    @pyqtSlot(name='on_close_path_clicked')
    def __closePath(self):
        """Add a the last point same as first point."""
        currentPath = self.currentPath()
        if (self.path_list.count() > 1) and (currentPath[0] != currentPath[-1]):
            self.addPoint(*currentPath[0])
    
    @pyqtSlot(name='on_point_up_clicked')
    def __moveUpPoint(self):
        """Target point move up."""
        row = self.path_list.currentRow()
        if not ((row > 0) and (self.path_list.count() > 1)):
            return
        path = self.currentPath()
        path.insert(row - 1, (path[row][0], path[row][1]))
        del path[row + 1]
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row - 1, "({}, {})".format(x, y))
        self.path_list.takeItem(row + 1)
        self.path_list.setCurrentRow(row - 1)
        self.__currentPathChanged()
    
    @pyqtSlot(name='on_point_down_clicked')
    def __moveDownPoint(self):
        """Target point move down."""
        row = self.path_list.currentRow()
        if not (
            (row < self.path_list.count() - 1) and
            (self.path_list.count() > 1)
        ):
            return
        path = self.currentPath()
        path.insert(row + 2, (path[row][0], path[row][1]))
        del path[row]
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row+2, "({}, {})".format(x, y))
        self.path_list.takeItem(row)
        self.path_list.setCurrentRow(row+1)
        self.__currentPathChanged()
    
    @pyqtSlot(name='on_point_delete_clicked')
    def __deletePoint(self):
        """Delete a target point."""
        row = self.path_list.currentRow()
        if not row > -1:
            return
        del self.currentPath()[row]
        self.path_list.takeItem(row)
        self.__currentPathChanged()
    
    def __ableToGenerate(self):
        """Set button enable if all the data are already."""
        self.pointNum.setText(
            "<p><span style=\"font-size:12pt;"
            "color:#00aa00;\">{}</span></p>".format(self.path_list.count())
        )
        n = (
            bool(self.mech_params) and
            (self.path_list.count() > 1) and
            bool(self.Expression.text())
        )
        self.path_adjust_button.setEnabled(n)
        self.synthesis_button.setEnabled(n)
    
    @pyqtSlot(name='on_synthesis_button_clicked')
    def __synthesis(self):
        """Start synthesis."""
        #Check if the number of target points are same.
        leng = -1
        for path in self.path.values():
            if leng<0:
                leng = len(path)
            if len(path)!=leng:
                QMessageBox.warning(self,
                    "Target Error",
                    "The length of target paths should be the same."
                )
                return
        #Get the algorithm type.
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        elif self.type2.isChecked():
            type_num = AlgorithmType.DE
        #Deep copy it so the pointer will not the same.
        mech_params = deepcopy(self.mech_params)
        mech_params['Target'] = deepcopy(self.path)
        
        def name_in_table(name: str) -> int:
            """Find a name and return the row from the table."""
            for row in range(self.parameter_list.rowCount()):
                if self.parameter_list.item(row, 0).text() == name:
                    return row
        
        for key in ('Driver', 'Follower'):
            for name in mech_params[key]:
                row = name_in_table(name)
                mech_params[key][name] = (
                    self.parameter_list.cellWidget(row, 2).value(),
                    self.parameter_list.cellWidget(row, 3).value(),
                    self.parameter_list.cellWidget(row, 4).value(),
                )
        
        #Start progress dialog.
        dlg = ProgressDialog(type_num, mech_params, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            return
        for m in dlg.mechanisms:
            self.__mechanism_data.append(m)
            self.__addResult(m)
        self.__setTime(dlg.time_spand)
        self.unsaveFunc()
        QMessageBox.information(self,
            "Dimensional Synthesis",
            "Your tasks is all completed.",
            QMessageBox.Ok
        )
        print("Finished.")
    
    def __setTime(self, time: float):
        """Set the time label."""
        self.timeShow.setText(
            "<html><head/><body><p><span style=\"font-size:16pt\">" +
            "{}[min] {:.02f}[s]".format(int(time // 60), time % 60) +
            "</span></p></body></html>"
        )
    
    def __addResult(self, result: Dict[str, Any]):
        """Add result items, except add to the list."""
        item = QListWidgetItem(result['Algorithm'])
        interrupt = result['interrupted']
        if interrupt=='False':
            item.setIcon(QIcon(QPixmap(":/icons/task-completed.png")))
        elif interrupt=='N/A':
            item.setIcon(QIcon(QPixmap(":/icons/question-mark.png")))
        else:
            item.setIcon(QIcon(QPixmap(":/icons/interrupted.png")))
        text = "{} ({})".format(
            result['Algorithm'],
            "No interrupt." if interrupt=='False' else "Interrupt at {}".format(interrupt)
        )
        if interrupt == 'N/A':
            text += "\n※Completeness is not clear."
        item.setToolTip(text)
        self.result_list.addItem(item)
    
    @pyqtSlot(name='on_delete_button_clicked')
    def __deleteResult(self):
        """Delete a result."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        reply = QMessageBox.question(self,
            "Delete",
            "Delete this result from list?"
        )
        if reply != QMessageBox.Yes:
            return
        del self.__mechanism_data[row]
        self.result_list.takeItem(row)
        self.unsaveFunc()
        self.__hasResult()
    
    @pyqtSlot(name='on_result_list_clicked')
    def __hasResult(self):
        """Set enable if there has any result."""
        for button in [
            self.merge_button,
            self.delete_button,
            self.result_load_settings,
            self.result_chart,
            self.result_clipboard
        ]:
            button.setEnabled(self.result_list.currentRow()>-1)
    
    @pyqtSlot(QModelIndex, name='on_result_list_doubleClicked')
    def __showResult(self, index: QModelIndex):
        """Double click result item can show up preview dialog."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        dlg = PreviewDialog(self.__mechanism_data[row], self.__getPath(row), self)
        dlg.show()
        dlg.exec_()
    
    @pyqtSlot(name='on_merge_button_clicked')
    def __mergeResult(self):
        """Merge mechanism into main canvas."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        reply = QMessageBox.question(self,
            "Merge",
            "Merge this result to your canvas?"
        )
        if reply == QMessageBox.Yes:
            self.mergeResult(row, self.__getPath(row))
    
    def __getPath(self, row: int):
        """Using result data to generate paths of mechanism."""
        Result = self.__mechanism_data[row]
        exprs = []
        for expr in Result['Expression'].split(';'):
            func = strbefore(expr, '[')
            params = strbetween(expr, '[', ']').split(',')
            target = strbetween(expr, '(', ')')
            params.insert(0, func)
            params.append(target)
            exprs.append(tuple(params))
        pos = {}
        for name in Result['pos']:
            try:
                pos[name] = Result['P{}'.format(name)]
            except KeyError:
                pos[name] = Result['pos'][name]
        
        vpoints = graph2vpoints(
            Graph(Result['Graph']),
            pos,
            Result['cus'],
            Result['same']
        )
        vpoint_count = len(vpoints)
        
        path = []
        for i in range(vpoint_count):
            path.append([])
        
        #Cumulative angle
        i_count = sum(1 for e in exprs if e[0] == 'PLAP')
        angles_cum = [0.] * i_count
        nan = float('nan')
        for interval in (3, -3):
            #Driver pointer
            dp = 0
            angles = [0.] * i_count
            while dp < i_count:
                try:
                    result = expr_solving(
                        exprs,
                        {n: 'P{}'.format(n) for n in range(len(vpoints))},
                        vpoints,
                        angles
                    )
                except Exception:
                    #Update with error sign.
                    for i in range(vpoint_count):
                        path[i].append((nan, nan))
                    #Back to last feasible solution.
                    angles[dp] -= interval
                    dp += 1
                else:
                    #Update with result.
                    for i in range(vpoint_count):
                        if result[i][0] == tuple:
                            path[i].append(result[i][1])
                            vpoints[i].move(result[i][0], result[i][1])
                        else:
                            path[i].append(result[i])
                            vpoints[i].move(result[i])
                    angles[dp] += interval
                    angles[dp] %= 360
                    angles_cum[dp] += abs(interval)
                    if angles_cum[dp] > 360:
                        angles[dp] -= interval
                        dp += 1
        return path
    
    @pyqtSlot(name='on_result_chart_clicked')
    def __showResultChart(self):
        """Show up the chart dialog."""
        dlg = ChartDialog("Convergence Value", self.__mechanism_data, self)
        dlg.show()
        dlg.exec_()
    
    @pyqtSlot(name='on_result_clipboard_clicked')
    def __copyResultText(self):
        """Copy pretty print result as text."""
        QApplication.clipboard().setText(
            pprint.pformat(self.__mechanism_data[self.result_list.currentRow()])
        )
    
    @pyqtSlot(name='on_save_profile_clicked')
    def __saveProfile(self):
        """Save as new profile to collection widget."""
        if not self.mech_params:
            return
        name, ok = QInputDialog.getText(self,
            "Profile name",
            "Please enter the profile name:"
        )
        if not ok:
            return
        i = 0
        while (name not in self.collections) and (not name):
            name = "Structure_{}".format(i)
        
        mech_params = deepcopy(self.mech_params)
        for key in ['Driver', 'Follower', 'Target']:
            for mp in mech_params[key]:
                mech_params[key][mp] = None
        
        self.collections[name] = mech_params
        self.unsaveFunc()
    
    @pyqtSlot(name='on_load_profile_clicked')
    def __loadProfile(self):
        """Load profile from collections dialog."""
        dlg = CollectionsDialog(
            self.collections,
            self.getCollection,
            self
        )
        dlg.show()
        if dlg.exec_():
            self.__setProfile(dlg.name(), dlg.params())
    
    def __setProfile(self, profile_name: str, params: Dict[str, Any]):
        """Set profile to sub-widgets."""
        self.__clearSettings()
        self.profile_name.setText(profile_name)
        self.mech_params = deepcopy(params)
        self.Expression.setText(self.mech_params['Expression'])
        self.Link_expr.setText(self.mech_params['Link_expr'])
        for name in sorted(self.mech_params['Target']):
            self.target_points.addItem(name)
            path = self.mech_params['Target'][name]
            if path:
                self.path[name] = path.copy()
            else:
                self.path[name] = []
        if self.target_points.count():
            self.target_points.setCurrentRow(0)
        
        gj = {}
        for key in ('Driver', 'Follower'):
            gj.update(self.mech_params[key])
        
        link_list = set()
        angle_list = set()
        for expr in self.mech_params['Expression'].split(';'):
            for e in strbetween(expr, '[', ']').split(','):
                if e.startswith('L'):
                    link_list.add(e)
                if e.startswith('a'):
                    angle_list.add(e)
        link_count = len(link_list)
        angle_count = len(angle_list)
        
        self.parameter_list.setRowCount(0)
        self.parameter_list.setRowCount(len(gj) + link_count + angle_count)
        
        def spinbox(
            v: float,
            *,
            minimum: float = 0,
            maximum: float = 1000000.0,
            prefix: bool = False
        ) -> QDoubleSpinBox:
            s = QDoubleSpinBox()
            s.setMinimum(minimum)
            s.setMaximum(maximum)
            s.setSingleStep(10.0)
            s.setValue(v)
            if prefix:
                s.setPrefix("±")
            return s
        
        row = 0
        for name in sorted(gj):
            coord = gj[name]
            self.parameter_list.setItem(row, 0, QTableWidgetItem(name))
            if name in self.mech_params['Driver']:
                role = 'Driver'
            else:
                role = 'Follower'
            self.parameter_list.setItem(row, 1, QTableWidgetItem(role))
            x, y = self.mech_params['pos'][int(name.replace('P', ''))]
            s1 = spinbox(coord[0] if coord else x, minimum=-1000000.0)
            s2 = spinbox(coord[1] if coord else y, minimum=-1000000.0)
            s3 = spinbox(coord[2] if coord else 50., prefix=True)
            self.parameter_list.setCellWidget(row, 2, s1)
            self.parameter_list.setCellWidget(row, 3, s2)
            self.parameter_list.setCellWidget(row, 4, s3)
            #Signal connections.
            for s in (s1, s2, s3):
                s.valueChanged.connect(self.updateRange)
            row += 1
        
        def set_by_center(
            i: int,
            get_range: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by center."""
            
            @pyqtSlot(float)
            def func(value: float):
                half_range = get_range() / 2
                self.mech_params['upper'][i] = value + half_range
                self.mech_params['lower'][i] = value - half_range
            
            return func
        
        def set_by_range(
            i: int,
            get_value: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by range."""
            
            @pyqtSlot(float)
            def func(value: float):
                center = get_value()
                half_range = value / 2
                self.mech_params['upper'][i] = center + half_range
                self.mech_params['lower'][i] = center - half_range
            
            return func
        
        for name in ('upper', 'lower'):
            if name not in self.mech_params:
                self.mech_params[name] = [0.] * (link_count + angle_count)
        
        for i, name in enumerate(sorted(link_list) + sorted(angle_list)):
            name_item = QTableWidgetItem(name)
            name_item.setToolTip(name)
            self.parameter_list.setItem(row, 0, name_item)
            self.parameter_list.setItem(row, 1, QTableWidgetItem('Limit'))
            #Set values (it will be same if not in the 'mech_params').
            upper = self.mech_params['upper'][i]
            if upper == 0:
                upper = 100. if name in link_list else 360.
            lower = self.mech_params['lower'][i]
            if lower == 0 and name in link_list:
                lower = 5.
            self.mech_params['upper'][i] = upper
            self.mech_params['lower'][i] = lower
            #Spin box.
            error_range = upper - lower
            default_value = error_range / 2 + lower
            if name in link_list:
                s1 = spinbox(default_value)
            else:
                s1 = spinbox(default_value, maximum=360.)
            self.parameter_list.setCellWidget(row, 2, s1)
            s2 = spinbox(error_range, prefix=True)
            self.parameter_list.setCellWidget(row, 4, s2)
            #Signal connections.
            s1.valueChanged.connect(set_by_center(i, s2.value))
            s2.valueChanged.connect(set_by_range(i, s1.value))
            row += 1
        
        self.PreviewCanvas.from_profile(self.mech_params)
        self.updateRange()
        self.__ableToGenerate()
        if not self.Expression.text():
            QMessageBox.warning(self,
                "Profile cannot use",
                "This profile has no any solutions, " +
                "you can set it in the \"Triangular iteration\" page."
            )
    
    @pyqtSlot(name='on_result_load_settings_clicked')
    def __loadResultSettings(self):
        """Load settings from a result."""
        self.__hasResult()
        row = self.result_list.currentRow()
        if not row > -1:
            return
        self.__clearSettings()
        Result = self.__mechanism_data[row]
        if Result['Algorithm'] == str(AlgorithmType.RGA):
            self.type0.setChecked(True)
        elif Result['Algorithm'] == str(AlgorithmType.Firefly):
            self.type1.setChecked(True)
        elif Result['Algorithm'] == str(AlgorithmType.DE):
            self.type2.setChecked(True)
        #Copy to mechanism params.
        self.__setProfile("External setting", Result)
        self.__setTime(Result['time'])
        #Load settings.
        self.alg_options.clear()
        self.alg_options.update(Result['settings'])
    
    @pyqtSlot(name='on_type0_clicked')
    @pyqtSlot(name='on_type1_clicked')
    @pyqtSlot(name='on_type2_clicked')
    def __setAlgorithmToDefault(self):
        """Set the algorithm settings to default."""
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        if self.type0.isChecked():
            self.alg_options.update(GeneticPrams)
        elif self.type1.isChecked():
            self.alg_options.update(FireflyPrams)
        elif self.type2.isChecked():
            self.alg_options.update(DifferentialPrams)
    
    @pyqtSlot(name='on_advance_button_clicked')
    def __showAdvance(self):
        """Get the settings from advance dialog."""
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        elif self.type2.isChecked():
            type_num = AlgorithmType.DE
        dlg = AlgorithmOptionDialog(type_num, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            return
        self.alg_options['report'] = dlg.report.value()
        if dlg.maxGen_option.isChecked():
            self.alg_options['maxGen'] = dlg.maxGen.value()
        elif dlg.minFit_option.isChecked():
            self.alg_options['minFit'] = dlg.minFit.value()
        elif dlg.maxTime_option.isChecked():
            #Three spinbox value translate to second.
            self.alg_options['maxTime'] = (
                dlg.maxTime_h.value() * 3600 +
                dlg.maxTime_m.value() * 60 +
                dlg.maxTime_s.value()
            )
        
        def from_table(row: int) -> Union[int, float]:
            """Get algorithm data from table."""
            return dlg.alg_table.cellWidget(row, 1).value()
        
        pop_size = dlg.pop_size.value()
        if type_num == AlgorithmType.RGA:
            self.alg_options['nPop'] = pop_size
            for i, tag in enumerate(('pCross', 'pMute', 'pWin', 'bDelta')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.Firefly:
            self.alg_options['n'] = pop_size
            for i, tag in enumerate(('alpha', 'betaMin', 'gamma', 'beta0')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.DE:
            self.alg_options['NP'] = pop_size
            for i, tag in enumerate(('strategy', 'F', 'CR')):
                self.alg_options[tag] = from_table(i)
    
    @pyqtSlot(float)
    def updateRange(self, p0: Optional[float] = None):
        """Update range values to main canvas."""
        
        def t(x: int, y: int):
            item = self.parameter_list.item(x, y)
            if item:
                return item.text()
            else:
                return self.parameter_list.cellWidget(x, y).value()
        
        self.updateRanges({
            t(row, 0): (t(row, 2), t(row, 3), t(row, 4))
            for row in range(self.parameter_list.rowCount())
            if t(row, 1) in ('Follower', 'Driver')
        })
    
    @pyqtSlot(name='on_expr_copy_clicked')
    def __copyExpr(self):
        """Copy profile expression."""
        text = self.Expression.text()
        if text:
            QApplication.clipboard().setText(text)
    
    @pyqtSlot(name='on_link_expr_copy_clicked')
    def __copyLinkExpr(self):
        """Copy profile link expression."""
        text = self.Link_expr.text()
        if text:
            QApplication.clipboard().setText(text)
