# -*- coding: utf-8 -*-
##Pyslvs - Open Source Planar Linkage Mechanism Simulation and Dimensional Synthesis System.
##Copyright (C) 2016-2017 Yuan Chang
##E-mail: pyslvs@gmail.com
##
##This program is free software; you can redistribute it and/or modify
##it under the terms of the GNU Affero General Public License as published by
##the Free Software Foundation; either version 3 of the License, or
##(at your option) any later version.
##
##This program is distributed in the hope that it will be useful,
##but WITHOUT ANY WARRANTY; without even the implied warranty of
##MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
##GNU Affero General Public License for more details.
##
##You should have received a copy of the GNU Affero General Public License
##along with this program; if not, write to the Free Software
##Foundation, Inc., 59 Temple Place - Suite 330, Boston, MA 02111-1307, USA.

from ...QtModules import *
from .Ui_Path_Solving import Ui_Form as PathSolving_Form
from ...graphics.ChartGraphics import ChartDialog
from ...graphics.Path_Solving_preview import PreviewDialog
from ...kernel.pyslvs_algorithm.TS import solver, Direction
from .Path_Solving_options import Path_Solving_options_show
from .Path_Solving_path_adjust import Path_Solving_path_adjust_show
from .Path_Solving_progress import Path_Solving_progress_show
from .Path_Solving_series import Path_Solving_series_show
import csv
import openpyxl
import re
import platform

#SystemTrayIcon
class progress_systemTrayIcon(QSystemTrayIcon):
    def __init__(self, parent=None):
        QSystemTrayIcon.__init__(self, QIcon(QPixmap(":/icons/main_big.png")), parent)
        self.menu = QMenu(parent)
        self.setContextMenu(self.menu)
        self.messageClicked.connect(self.showMainWindow)
        self.dlg = None
        if platform.system().lower()=='windows':
            self.setToolTip("Pyslvs\nClick here to minimize / show Pyslvs.")
        else:
            self.setToolTip("<html><head/><body><h2>Pyslvs</h2><p>Click here to minimize / show Pyslvs.</p></body></html>")
        self.activated.connect(self.clicked)
        self.mainState = parent.windowState()
    
    def setDialog(self, dlg):
        self.dlg = dlg
    
    @pyqtSlot()
    def showMainWindow(self):
        mainWindow = self.parent()
        mainWindow.setWindowState(self.mainState)
        mainWindow.activateWindow()
    
    @pyqtSlot(QSystemTrayIcon.ActivationReason)
    def clicked(self, ActivationReason):
        if ActivationReason==QSystemTrayIcon.Trigger:
            mainWindow = self.parent()
            if mainWindow.windowState()==Qt.WindowMinimized:
                mainWindow.setWindowState(self.mainState)
                if self.dlg:
                    self.dlg.setWindowState(self.mainState)
            else:
                self.mainState = mainWindow.windowState()
                mainWindow.showMinimized()
                if self.dlg:
                    self.dlg.showMinimized()

class Path_Solving_show(QWidget, PathSolving_Form):
    fixPointRange = pyqtSignal(tuple, float, tuple, float)
    addPathPoint = pyqtSignal(float, float)
    deletePathPoint = pyqtSignal(int)
    moveupPathPoint = pyqtSignal(int)
    movedownPathPoint = pyqtSignal(int)
    mergeResult = pyqtSignal(int, float, float, list, dict)
    GeneticPrams = {'nPop':500, 'pCross':0.95, 'pMute':0.05, 'pWin':0.95, 'bDelta':5.}
    FireflyPrams = {'n':80, 'alpha':0.01, 'betaMin':0.2, 'gamma':1., 'beta0':1.}
    DifferentialPrams = {'strategy':1, 'NP':400, 'F':0.6, 'CR':0.9}
    defaultSettings = {'maxGen':1500, 'report':1, 'IMin':5., 'LMin':5., 'FMin':5., 'AMin':0.,
        'IMax':300., 'LMax':300., 'FMax':300., 'AMax':360., 'algorithmPrams':DifferentialPrams}
    mechanismParams_4Bar = { #No 'targetPath'
        'Driving':'A',
        'Follower':'D',
        'Link':'L0,L1,L2,L3,L4',
        'Target':'E',
        'ExpressionName':'PLAP,PLLP,PLLP',
        'Expression':'A,L0,a0,D,B,B,L1,L2,D,C,B,L3,L4,C,E',
        'constraint':[{'driver':'L0', 'follower':'L2', 'connect':'L1'}],
        'formula':['PLAP','PLLP']}
    mechanismParams_4Bar['VARS'] = len(set(mechanismParams_4Bar['Expression'].split(',')))-2
    mechanismParams_8Bar = { #No 'targetPath'
        'Driving':'A',
        'Follower':'B',
        'Link':'L0,L1,L2,L3,L4,L5,L6,L7,L8,L9,L10',
        'Target':'H',
        'ExpressionName':'PLAP,PLLP,PLLP,PLLP,PLLP,PLLP',
        'Expression':'A,L0,a0,B,C,B,L2,L1,C,D,B,L4,L3,D,E,C,L5,L6,B,F,F,L8,L7,E,G,F,L9,L10,G,H',
        'constraint':[{'driver':'L0', 'follower':'L2', 'connect':'L1'}],
        'formula':['PLAP','PLLP']}
    mechanismParams_8Bar['VARS'] = len(set(mechanismParams_8Bar['Expression'].split(',')))-2
    
    def __init__(self, path, mechanism_data, env, unsave_func, parent=None):
        super(Path_Solving_show, self).__init__(parent)
        self.setupUi(self)
        #System Tray Icon Menu
        self.trayIcon = progress_systemTrayIcon(parent)
        self.path = path
        self.mechanism_data = mechanism_data
        self.env = env
        self.unsave_func = unsave_func
        for e in path:
            self.Point_list.addItem("({}, {})".format(e['x'], e['y']))
        for e in mechanism_data:
            self.addResult(e)
        self.Settings = self.defaultSettings
        self.algorithmPrams_default()
        self.Point_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.Point_list.customContextMenuRequested.connect(self.on_Point_list_context_menu)
        self.popMenu_list = QMenu(self)
        self.action_paste_from_clipboard = QAction("&Paste from clipboard", self)
        self.popMenu_list.addAction(self.action_paste_from_clipboard)
        self.Ar.valueChanged.connect(self.updateRange)
        self.Ax.valueChanged.connect(self.updateRange)
        self.Ay.valueChanged.connect(self.updateRange)
        self.Dr.valueChanged.connect(self.updateRange)
        self.Dx.valueChanged.connect(self.updateRange)
        self.Dy.valueChanged.connect(self.updateRange)
        self.isGenerate()
        self.isGetResult()
    
    @pyqtSlot()
    def on_clearAll_clicked(self):
        self.Point_list.setCurrentRow(0)
        for i in reversed(range(self.Point_list.count()+1)):
            self.on_remove_clicked()
        self.isGenerate()
    
    @pyqtSlot()
    def on_series_clicked(self):
        dlg = Path_Solving_series_show(self)
        dlg.show()
        if dlg.exec_():
            for e in dlg.path:
                self.on_add_clicked(e[0], e[1])
    
    def on_Point_list_context_menu(self, point):
        action = self.popMenu_list.exec_(self.Point_list.mapToGlobal(point))
        if action==self.action_paste_from_clipboard:
            data = QApplication.clipboard().text()
            data = re.split(',\t|\n', data)
            try:
                data = [(round(float(data[i]), 4), round(float(data[i+1]), 4)) for i in range(0, len(data), 2)]
                for e in data:
                    self.on_add_clicked(e[0], e[1])
            except:
                dlgbox = QMessageBox(QMessageBox.Warning, "File error", "Wrong format.\nIt should be look like this:"+
                    "\n0.0,[\\tab]0.0[\\n]"*3, (QMessageBox.Ok), self)
                if dlgbox.exec_():
                    pass
    
    @pyqtSlot()
    def on_importCSV_clicked(self):
        fileName, _ = QFileDialog.getOpenFileName(self, 'Open file...', self.env, "Text File(*.txt);;CSV File(*.csv)")
        if fileName:
            data = list()
            with open(fileName, newline=str()) as stream:
                reader = csv.reader(stream, delimiter=' ', quotechar='|')
                for row in reader:
                    data += ' '.join(row).split(',\t')
            try:
                data = [(round(float(data[i]), 4), round(float(data[i+1]), 4)) for i in range(0, len(data), 2)]
                for e in data:
                    self.on_add_clicked(e[0], e[1])
            except:
                dlgbox = QMessageBox(QMessageBox.Warning, "File error", "Wrong format.\nIt should be look like this:"+
                    "\n0.0,[\\tab]0.0[\\n]"*3, (QMessageBox.Ok), self)
                if dlgbox.exec_():
                    pass
    
    @pyqtSlot()
    def on_importXLSX_clicked(self):
        fileName, _ = QFileDialog.getOpenFileName(self, 'Open file...', self.env, "Microsoft Office Excel(*.xlsx *.xlsm *.xltx *.xltm)")
        if fileName:
            wb = openpyxl.load_workbook(fileName)
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            data = list()
            i = 1
            while True:
                x = ws.cell(row=i, column=1).value
                y = ws.cell(row=i, column=2).value
                if x==None or y==None:
                    break
                try:
                    data.append((round(float(x), 4), round(float(y), 4)))
                except:
                    dlgbox = QMessageBox(QMessageBox.Warning, "File error", "Wrong format.\nThe datasheet seems to including non-digital cell.", (QMessageBox.Ok), self)
                    if dlgbox.exec_():
                        break
                i += 1
            for e in data:
                self.on_add_clicked(e[0], e[1])
    
    @pyqtSlot()
    def on_pathAdjust_clicked(self):
        dlg = Path_Solving_path_adjust_show(self.path)
        dlg.show()
        if dlg.exec_():
            self.on_clearAll_clicked()
            for e in dlg.get_path():
                self.on_add_clicked(e[0], e[1])
    
    @pyqtSlot()
    def on_moveUp_clicked(self):
        n = self.Point_list.currentRow()
        if n>0 and self.Point_list.count()>1:
            self.moveupPathPoint.emit(n)
            x = self.Point_list.currentItem().text()[1:-1].split(', ')[0]
            y = self.Point_list.currentItem().text()[1:-1].split(', ')[1]
            self.Point_list.insertItem(n-1, '({}, {})'.format(x, y))
            self.Point_list.takeItem(n+1)
            self.Point_list.setCurrentRow(n-1)
    
    @pyqtSlot()
    def on_moveDown_clicked(self):
        n = self.Point_list.currentRow()
        if n<self.Point_list.count()-1 and self.Point_list.count()>1:
            self.movedownPathPoint.emit(n)
            c = self.Point_list.currentItem().text()[1:-1].split(', ')
            self.Point_list.insertItem(n+2, "({}, {})".format(c[0], c[1]))
            self.Point_list.takeItem(n)
            self.Point_list.setCurrentRow(n+1)
    
    def addPath(self, x, y):
        self.Point_list.addItem('({}, {})'.format(x, y))
        self.isGenerate()
    
    @pyqtSlot()
    def on_add_clicked(self, x=False, y=False):
        if x is False:
            x = self.X_coordinate.value()
            y = self.Y_coordinate.value()
        self.addPathPoint.emit(x, y)
        self.Point_list.addItem("({}, {})".format(x, y))
        self.isGenerate()
    @pyqtSlot()
    def on_remove_clicked(self):
        if self.Point_list.currentRow()>-1:
            self.deletePathPoint.emit(self.Point_list.currentRow())
            self.Point_list.takeItem(self.Point_list.currentRow())
            self.isGenerate()
    
    def isGenerate(self):
        self.pointNum.setText(
            "<html><head/><body><p><span style=\"font-size:12pt; color:#00aa00;\">{}</span></p></body></html>".format(self.Point_list.count()))
        n = self.Point_list.count()>1
        self.pathAdjust.setEnabled(n)
        self.GenerateLocal.setEnabled(n)
        self.GenerateZMQ.setEnabled(n)
    
    @pyqtSlot()
    def on_GenerateLocal_clicked(self):
        self.startAlgorithm()
    @pyqtSlot()
    def on_GenerateZMQ_clicked(self):
        self.startAlgorithm(hasPort=True)
    def startAlgorithm(self, hasPort=False):
        self.trayIcon.show()
        type_num, mechanismParams, generateData = self.getGenerate()
        dlg = Path_Solving_progress_show(type_num, mechanismParams, generateData, self.Settings['algorithmPrams'],
            PORT=self.portText.text() if hasPort else None, parent=self.parent())
        self.trayIcon.setDialog(dlg)
        dlg.show()
        if dlg.exec_():
            self.mechanism_data += dlg.mechanisms
            for m in dlg.mechanisms:
                self.addResult(m)
            self.setTime(dlg.time_spand)
            self.unsave_func()
            if self.trayIcon.supportsMessages():
                self.trayIcon.showMessage("Algorithm completed!", "You can see the results in Pyslvs.")
            dlgbox = QMessageBox(QMessageBox.Information, "Dimensional Synthesis", "Your tasks is all completed.", (QMessageBox.Ok), self.parent())
            if dlgbox.exec_():
                print('Finished.')
        self.trayIcon.hide()
    def getGenerate(self):
        type_num = 0 if self.type0.isChecked() else 1 if self.type1.isChecked() else 2
        mechanismParams = self.mechanismParams_4Bar if self.FourBar.isChecked() else self.mechanismParams_8Bar
        link_q = mechanismParams['VARS']-7
        upper = [self.Ax.value()+self.Ar.value()/2, self.Ay.value()+self.Ar.value()/2, self.Dx.value()+self.Dr.value()/2, self.Dy.value()+self.Dr.value()/2,
            self.Settings['IMax'], self.Settings['LMax'], self.Settings['FMax']]+[self.Settings['LMax']]*link_q
        lower = [self.Ax.value()-self.Ar.value()/2, self.Ay.value()-self.Ar.value()/2, self.Dx.value()-self.Dr.value()/2, self.Dy.value()-self.Dr.value()/2,
            self.Settings['IMin'], self.Settings['LMin'], self.Settings['FMin']]+[self.Settings['LMin']]*link_q
        mechanismParams['targetPath'] = tuple((e['x'], e['y']) for e in self.path)
        p = len(self.path)
        generateData = {
            'nParm':p+mechanismParams['VARS'],
            'upper':upper+[self.Settings['AMax']]*p,
            'lower':lower+[self.Settings['AMin']]*p,
            'maxGen':self.Settings['maxGen'],
            'report':int(self.Settings['maxGen']*self.Settings['report']/100)}
        return type_num, mechanismParams, generateData
    
    def setTime(self, time_spand):
        sec = round(time_spand%60, 2)
        mins = int(time_spand/60)
        self.timeShow.setText("<html><head/><body><p><span style=\"font-size:12pt\">{}[min] {}[s]</span></p></body></html>".format(mins, sec))
    
    def addResult(self, result):
        item = QListWidgetItem("{} ({} gen)".format(result['Algorithm'], result['generateData']['maxGen']))
        interrupt = result['interruptedGeneration']
        if interrupt=='False':
            item.setIcon(QIcon(QPixmap(":/icons/task-completed.png")))
        elif interrupt=='N/A':
            item.setIcon(QIcon(QPixmap(":/icons/question-mark.png")))
        else:
            item.setIcon(QIcon(QPixmap(":/icons/interrupted.png")))
        keys = sorted(list(result.keys()))
        info = (["{}: {}".format(k, result[k]) for k in keys if 'x' in k or 'y' in k or 'L' in k]+
            ["\nClick to apply setting."]+["Double click to see dynamic preview."])
        item.setToolTip('\n'.join(["[{}] ({}{} gen)".format(result['Algorithm'],
            '' if interrupt=='False' else interrupt+'-', result['generateData']['maxGen'])]+
            (["※ Completeness is not clear." if interrupt=='N/A' else ''])+info))
        self.Result_list.addItem(item)
    
    @pyqtSlot()
    def on_deleteButton_clicked(self):
        row = self.Result_list.currentRow()
        del self.mechanism_data[row]
        self.Result_list.takeItem(row)
        self.unsave_func()
        self.isGetResult()
    
    def isGetResult(self):
        for button in [self.mergeButton, self.deleteButton]:
            button.setEnabled(self.Result_list.currentRow()>-1)
    
    @pyqtSlot(QModelIndex)
    def on_Result_list_doubleClicked(self, index):
        row = self.Result_list.currentRow()
        if row!=-1:
            mechanism = self.mechanism_data[row]
            _, _, _, _, Paths = self.legal_crank(row)
            dlg = PreviewDialog(mechanism, Paths, self)
            dlg.show()
            if dlg.exec_():
                pass
    
    @pyqtSlot()
    def on_mergeButton_clicked(self):
        reply = QMessageBox.question(self, 'Prompt Message', "Merge this result to your canvas?",
            (QMessageBox.Apply | QMessageBox.Cancel), QMessageBox.Apply)
        if reply==QMessageBox.Apply:
            self.mergeResult.emit(*self.legal_crank(self.Result_list.currentRow()))
    
    def legal_crank(self, row):
        Result = self.mechanism_data[row]
        path = Result['mechanismParams']['targetPath']
        pointAvg = sum([e[1] for e in path])/len(path)
        other = (Result['Ay']+Result['Dy'])/2>pointAvg and Result['Ax']<Result['Dx']
        answer = [False]
        startAngle = False
        endAngle = False
        expression = Result['mechanismParams']['Expression'].split(',')
        expression_tag = tuple(tuple(expression[i+j] for j in range(5)) for i in range(0, len(expression), 5))
        expression_result = [exp[-1] for exp in expression_tag]
        Paths = {tag:list() for tag in expression_result}
        for a in range(360+1):
            Directions = [Direction(p1=(Result['Ax'], Result['Ay']), p2=(Result['Ax']+10, Result['Ay']), len1=Result['L0'], angle=a, other=other)]
            for exp in expression_tag[1:]:
                p1 = (Result['Ax'], Result['Ay']) if exp[0]=='A' else expression_result.index(exp[0]) if exp[0] in expression_result else (Result['Dx'], Result['Dy'])
                p2 = (Result['Ax'], Result['Ay']) if exp[3]=='A' else expression_result.index(exp[3]) if exp[3] in expression_result else (Result['Dx'], Result['Dy'])
                Directions.append(Direction(p1=p1, p2=p2, len1=Result[exp[1]], len2=Result[exp[2]], other=other))
            s = solver(Directions) #showError=(a==0)
            s_answer = s.answer()
            answerT = [(Result['Ax'], Result['Ay']), (Result['Dx'], Result['Dy'])]+s_answer
            if not False in answerT:
                if startAngle is False:
                    startAngle = float(a)
                    answer = answerT
                endAngle = float(a)
            for i, a in enumerate(s_answer):
                Paths[expression_result[i]].append(a)
        return row, startAngle, endAngle, answer, Paths
    
    @pyqtSlot()
    def on_getTimeAndFitness_clicked(self):
        dlg = ChartDialog("Convergence Value", self.mechanism_data, self)
        dlg.show()
    
    @pyqtSlot(int)
    def on_Result_list_currentRowChanged(self, cr):
        self.isGetResult()
        if cr>-1 and cr!=len(self.mechanism_data):
            args = self.mechanism_data[cr]
            keys = set(args['algorithmPrams'].keys())
            if keys==set(self.GeneticPrams.keys()):
                self.type0.setChecked(True)
            elif keys==set(self.FireflyPrams.keys()):
                self.type1.setChecked(True)
            elif keys==set(self.DifferentialPrams.keys()):
                self.type2.setChecked(True)
            if args['mechanismParams']['Link']=='L0,L1,L2,L3,L4,L5,L6,L7,L8,L9,L10':
                self.EightBar.setChecked(True)
            else:
                self.FourBar.setChecked(True)
            self.setTime(args['time'])
            generateData = args['generateData']
            self.Ax.setValue((generateData['upper'][0]+generateData['lower'][0])/2)
            self.Ay.setValue((generateData['upper'][1]+generateData['lower'][1])/2)
            self.Ar.setValue(abs(generateData['upper'][0]-self.Ax.value())*2)
            self.Dx.setValue((generateData['upper'][2]+generateData['lower'][2])/2)
            self.Dy.setValue((generateData['upper'][3]+generateData['lower'][3])/2)
            self.Dr.setValue(abs(generateData['upper'][2]-self.Dx.value())*2)
            self.Settings = {'maxGen':generateData['maxGen'],
                'report':0 if generateData['report']==0 else generateData['maxGen']/generateData['report']/100,
                'IMax':generateData['upper'][4], 'IMin':generateData['lower'][4],
                'LMax':generateData['upper'][5], 'LMin':generateData['lower'][5],
                'FMax':generateData['upper'][6], 'FMin':generateData['lower'][6],
                'AMax':generateData['upper'][-1], 'AMin':generateData['lower'][-1]}
            self.Settings['algorithmPrams'] = args['algorithmPrams']
            self.on_clearAll_clicked()
            for e in args['mechanismParams']['targetPath']:
                self.on_add_clicked(e[0], e[1])
    
    def algorithmPrams_default(self):
        type_num = 0 if self.type0.isChecked() else 1 if self.type1.isChecked() else 2
        if type_num==0:
            self.Settings['algorithmPrams'] = self.GeneticPrams
        elif type_num==1:
            self.Settings['algorithmPrams'] = self.FireflyPrams
        elif type_num==2:
            self.Settings['algorithmPrams'] = self.DifferentialPrams
    @pyqtSlot(bool)
    def on_type0_toggled(self, checked):
        self.algorithmPrams_default()
    @pyqtSlot(bool)
    def on_type1_toggled(self, checked):
        self.algorithmPrams_default()
    @pyqtSlot(bool)
    def on_type2_toggled(self, checked):
        self.algorithmPrams_default()
    
    @pyqtSlot()
    def on_advanceButton_clicked(self):
        type_num = "Genetic Algorithm" if self.type0.isChecked() else "Firefly Algorithm" if self.type1.isChecked() else "Differential Evolution"
        dlg = Path_Solving_options_show(type_num, self.Settings)
        dlg.show()
        if dlg.exec_():
            tablePL = lambda row: dlg.PLTable.cellWidget(row, 1).value()
            self.Settings = {'maxGen':dlg.maxGen.value(), 'report':dlg.report.value(),
                'IMax':tablePL(0), 'IMin':tablePL(1),
                'LMax':tablePL(2), 'LMin':tablePL(3),
                'FMax':tablePL(4), 'FMin':tablePL(5),
                'AMax':tablePL(6), 'AMin':tablePL(7)}
            tableAP = lambda row: dlg.APTable.cellWidget(row, 1).value()
            popSize = dlg.popSize.value()
            if type_num=="Genetic Algorithm":
                self.Settings['algorithmPrams'] = {'nPop':popSize, 'pCross':tableAP(0), 'pMute':tableAP(1), 'pWin':tableAP(2), 'bDelta':tableAP(3)}
            elif type_num=="Firefly Algorithm":
                self.Settings['algorithmPrams'] = {'n':popSize, 'alpha':tableAP(0), 'betaMin':tableAP(1), 'gamma':tableAP(2), 'beta0':tableAP(3)}
            elif type_num=="Differential Evolution":
                self.Settings['algorithmPrams'] = {'NP':popSize, 'strategy':tableAP(0), 'F':tableAP(1), 'CR':tableAP(2)}
    
    @pyqtSlot(float)
    def updateRange(self, p0=None):
        self.fixPointRange.emit((self.Ax.value(), self.Ay.value()), self.Ar.value(), (self.Dx.value(), self.Dy.value()), self.Dr.value())
