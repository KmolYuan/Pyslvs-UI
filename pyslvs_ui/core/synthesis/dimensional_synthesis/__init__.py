# -*- coding: utf-8 -*-

"""'dimensional_synthesis' module contains
dimensional synthesis functional interfaces.
"""

from __future__ import annotations

__author__ = "Yuan Chang"
__copyright__ = "Copyright (C) 2016-2019"
__license__ = "AGPL"
__email__ = "pyslvs@gmail.com"

from typing import (
    TYPE_CHECKING,
    List,
    Tuple,
    Sequence,
    Dict,
    Callable,
    Union,
    Any,
)
from math import hypot
import csv
import pprint
from copy import deepcopy
from re import split as char_split
from openpyxl import load_workbook
from qtpy.QtCore import Slot, QModelIndex
from qtpy.QtWidgets import (
    QWidget,
    QApplication,
    QMessageBox,
    QHeaderView,
    QListWidgetItem,
    QInputDialog,
    QDoubleSpinBox,
    QTableWidgetItem,
)
from qtpy.QtGui import QIcon, QPixmap
from pyslvs import (
    vpoints_configure,
    expr_solving,
    parse_pos,
    parse_vpoints,
    parse_vlinks,
)
from pyslvs_ui.core.graphics import PreviewCanvas
from pyslvs_ui.core.synthesis import CollectionsDialog
from .dialogs import (
    GeneticPrams,
    FireflyPrams,
    defaultSettings,
    DifferentialPrams,
    AlgorithmType,
    AlgorithmOptionDialog,
    PathAdjustDialog,
    ProgressDialog,
    PreviewDialog,
    ChartDialog,
)
from .dimension_widget_ui import Ui_Form
if TYPE_CHECKING:
    from pyslvs_ui.core.widgets import MainWindowBase

__all__ = ['DimensionalSynthesis']
_Coord = Tuple[float, float]


class DimensionalSynthesis(QWidget, Ui_Form):

    """Dimensional synthesis widget.

    User can run the dimensional synthesis here.
    """

    def __init__(self, parent: MainWindowBase) -> None:
        """Reference names:

        + Iteration collections.
        + Result data.
        + Main window function references.
        """
        super(DimensionalSynthesis, self).__init__(parent)
        self.setupUi(self)

        self.mech_params: Dict[str, Any] = {}
        self.path: Dict[int, List[_Coord]] = {}

        # Some reference of 'collections'
        self.collections = parent.collections.configure_widget.collections
        self.get_collection = parent.get_configure
        self.input_from = parent.input_from
        self.workbook_no_save = parent.workbook_no_save
        self.merge_result = parent.merge_result
        self.update_ranges = parent.main_canvas.update_ranges
        self.set_solving_path = parent.main_canvas.set_solving_path
        self.prefer = parent.prefer

        # Data and functions
        self.mechanism_data: List[Dict[str, Any]] = []
        self.alg_options: Dict[str, Union[int, float]] = {}
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.__set_algorithm_default()

        self.preview_canvas = PreviewCanvas(self)
        self.preview_layout.addWidget(self.preview_canvas)

        # Splitter
        self.main_splitter.setStretchFactor(0, 100)
        self.main_splitter.setStretchFactor(1, 10)
        self.up_splitter.setSizes([80, 100])
        self.down_splitter.setSizes([20, 80])

        # Table widget column width
        header = self.parameter_list.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)

        self.clear()

    def clear(self) -> None:
        """Clear all sub-widgets."""
        self.mechanism_data.clear()
        self.result_list.clear()
        self.__clear_settings()
        self.__has_result()

    def __clear_settings(self) -> None:
        """Clear sub-widgets that contain the setting."""
        self.__clear_path(ask=False)
        self.path.clear()
        self.mech_params.clear()
        self.preview_canvas.clear()
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        self.alg_options.update(DifferentialPrams)
        self.profile_name.clear()
        self.type2.setChecked(True)
        self.parameter_list.setRowCount(0)
        self.target_points.clear()
        self.target_label.setVisible(self.has_target())
        self.expression_string.clear()
        self.update_range()
        self.__able_to_generate()

    def has_target(self) -> bool:
        """Return True if the panel is no target settings."""
        return self.target_points.count() > 0

    @Slot(name='on_clear_button_clicked')
    def __user_clear(self) -> None:
        if not self.profile_name.text():
            return

        if QMessageBox.question(
            self,
            "Clear setting",
            "Do you want to clear the setting?"
        ) == QMessageBox.Yes:
            self.__clear_settings()

    def load_results(self, mechanism_data: Sequence[Dict[str, Any]]) -> None:
        """Append results of workbook database to memory."""
        for e in mechanism_data:
            self.mechanism_data.append(e)
            self.__add_result(e)

    def __current_path_changed(self) -> None:
        """Call the canvas to update to current target path."""
        self.set_solving_path({
            f"P{name}": tuple(path) for name, path in self.path.items()
        })
        self.__able_to_generate()

    def current_path(self) -> List[_Coord]:
        """Return the pointer of current target path."""
        item = self.target_points.currentItem()
        if item is None:
            return []

        return self.path[int(item.text().replace('P', ''))]

    @Slot(str, name='on_target_points_currentTextChanged')
    def __set_target(self, _=None) -> None:
        """Switch to the current target path."""
        self.path_list.clear()
        for x, y in self.current_path():
            self.path_list.addItem(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @Slot(name='on_path_clear_clicked')
    def __clear_path(self, *, ask: bool = True) -> None:
        """Clear the current target path."""
        if ask:
            if QMessageBox.question(
                self,
                "Clear path",
                "Are you sure to clear the current path?"
            ) != QMessageBox.Yes:
                return

        self.current_path().clear()
        self.path_list.clear()
        self.__current_path_changed()

    @Slot(name='on_path_copy_clicked')
    def __copy_path(self) -> None:
        """Copy the current path coordinates to clipboard."""
        QApplication.clipboard().setText('\n'.join(
            f"{x},{y}" for x, y in self.current_path()
        ))

    @Slot(name='on_path_paste_clicked')
    def __paste_path(self) -> None:
        """Paste path data from clipboard."""
        self.__read_path_from_csv(char_split(r"[;,\n]", QApplication.clipboard().text()))

    @Slot(name='on_import_csv_button_clicked')
    def __import_csv(self) -> None:
        """Paste path data from a text file."""
        file_name = self.input_from(
            "path data",
            ["Text file (*.txt)", "Comma-Separated Values (*.csv)"]
        )
        if not file_name:
            return
        data = []
        with open(file_name, 'r', encoding='utf-8', newline='') as f:
            for row in csv.reader(f, delimiter=' ', quotechar='|'):
                data += " ".join(row).split(',')
        self.__read_path_from_csv(data)

    def __read_path_from_csv(self, raw_data: List[str]) -> None:
        """Turn string to float then add them to current target path."""
        try:
            data = [
                (round(float(raw_data[i]), 4), round(float(raw_data[i + 1]), 4))
                for i in range(0, len(raw_data), 2)
            ]
        except (IndexError, ValueError):
            QMessageBox.warning(
                self,
                "File error",
                "Wrong format.\nIt should be look like this:"
                + ("\n0.0,0.0[\\n]" * 3)
            )
        else:
            for x, y in data:
                self.add_point(x, y)

    @Slot(name='on_import_xlsx_button_clicked')
    def __import_xlsx(self) -> None:
        """Paste path data from a Excel file."""
        file_name = self.input_from(
            "Excel workbook",
            ["Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)"]
        )
        if not file_name:
            return
        wb = load_workbook(file_name)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        data = []
        # Keep finding until there is no value
        i = 1
        while True:
            x = ws.cell(row=i, column=1).value
            y = ws.cell(row=i, column=2).value
            if None in {x, y}:
                break
            try:
                data.append((round(float(x), 4), round(float(y), 4)))
            except (IndexError, AttributeError):
                QMessageBox.warning(
                    self,
                    "File error",
                    "Wrong format.\n"
                    "The data sheet seems to including non-digital cell."
                )
                break
            i += 1
        for x, y in data:
            self.add_point(x, y)

    @Slot(name='on_path_adjust_button_clicked')
    def __adjust_path(self) -> None:
        """Show up path adjust dialog and
        get back the changes of current target path.
        """
        dlg = PathAdjustDialog(self)
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return

        self.__clear_path(ask=False)
        for e in dlg.r_path:
            self.add_point(e[0], e[1])

        dlg.deleteLater()
        self.__current_path_changed()

    def add_point(self, x: float, y: float) -> None:
        """Add path data to list widget and current target path."""
        x = round(x, 4)
        y = round(y, 4)
        self.current_path().append((x, y))
        self.path_list.addItem(f"({x:.04f}, {y:.04f})")
        self.path_list.setCurrentRow(self.path_list.count() - 1)
        self.__current_path_changed()

    @Slot(float, float)
    def set_point(self, x: float, y: float) -> None:
        """Set the coordinate of current target path."""
        if not self.edit_target_point_button.isChecked():
            return
        for i, (cx, cy) in enumerate(self.current_path()):
            if hypot(x - cx, y - cy) < 5:
                index = i
                self.path_list.setCurrentRow(index)
                break
        else:
            return
        self.current_path()[index] = (x, y)
        self.path_list.item(index).setText(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @Slot(name='on_close_path_clicked')
    def __close_path(self) -> None:
        """Add a the last point same as first point."""
        current_path = self.current_path()
        if (self.path_list.count() > 1) and (current_path[0] != current_path[-1]):
            self.add_point(*current_path[0])

    @Slot(name='on_point_up_clicked')
    def __move_up_point(self) -> None:
        """Target point move up."""
        row = self.path_list.currentRow()
        if not ((row > 0) and (self.path_list.count() > 1)):
            return
        path = self.current_path()
        path.insert(row - 1, (path[row][0], path[row][1]))
        path.pop(row + 1)
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row - 1, f"({x}, {y})")
        self.path_list.takeItem(row + 1)
        self.path_list.setCurrentRow(row - 1)
        self.__current_path_changed()

    @Slot(name='on_point_down_clicked')
    def __move_down_point(self) -> None:
        """Target point move down."""
        row = self.path_list.currentRow()
        if not (
            (row < self.path_list.count() - 1)
            and (self.path_list.count() > 1)
        ):
            return
        path = self.current_path()
        path.insert(row + 2, (path[row][0], path[row][1]))
        path.pop(row)
        x, y = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row + 2, f"({x}, {y})")
        self.path_list.takeItem(row)
        self.path_list.setCurrentRow(row + 1)
        self.__current_path_changed()

    @Slot(name='on_point_delete_clicked')
    def __delete_point(self) -> None:
        """Delete a target point."""
        row = self.path_list.currentRow()
        if not row > -1:
            return
        self.current_path().pop(row)
        self.path_list.takeItem(row)
        self.__current_path_changed()

    def __able_to_generate(self) -> None:
        """Set button enable if all the data are already."""
        self.point_num.setText(
            "<p><span style=\"font-size:12pt;"
            f"color:#00aa00;\">{self.path_list.count()}</span></p>"
        )
        n = bool(self.mech_params and (self.path_list.count() > 1) and self.expression_string.text())
        self.path_adjust_button.setEnabled(n)
        self.synthesis_button.setEnabled(n)

    @Slot(name='on_synthesis_button_clicked')
    def __synthesis(self) -> None:
        """Start synthesis."""
        # Check if the number of target points are same.
        length = -1
        for path in self.path.values():
            if length < 0:
                length = len(path)
            if len(path) != length:
                QMessageBox.warning(
                    self,
                    "Target Error",
                    "The length of target paths should be the same."
                )
                return
        # Get the algorithm type
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        else:
            type_num = AlgorithmType.DE
        # Deep copy it so the pointer will not the same
        mech_params = deepcopy(self.mech_params)
        mech_params['Expression'] = parse_vpoints(mech_params.pop('Expression', []))
        mech_params['Target'] = deepcopy(self.path)

        def name_in_table(target_name: str) -> int:
            """Find a target_name and return the row from the table."""
            for r in range(self.parameter_list.rowCount()):
                if self.parameter_list.item(r, 0).text() == target_name:
                    return r
            return -1

        placement: Dict[int, Tuple[float, float, float]] = mech_params['Placement']
        for name in placement:
            row = name_in_table(f"P{name}")
            placement[name] = (
                self.parameter_list.cellWidget(row, 2).value(),
                self.parameter_list.cellWidget(row, 3).value(),
                self.parameter_list.cellWidget(row, 4).value(),
            )

        # Start progress dialog
        dlg = ProgressDialog(type_num, mech_params, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return

        mechanisms = dlg.mechanisms
        mechanisms_plot: List[Dict[str, Any]] = []
        for data in mechanisms:
            mechanisms_plot.append({
                'time_fitness': data.pop('time_fitness'),
                'Algorithm': data['Algorithm'],
            })
            self.mechanism_data.append(data)
            self.__add_result(data)
        self.__set_time(dlg.time_spend)
        self.workbook_no_save()
        dlg.deleteLater()

        dlg = ChartDialog("Convergence Data", mechanisms_plot, self)
        dlg.show()
        dlg.exec_()
        dlg.deleteLater()

    def __set_time(self, time: float) -> None:
        """Set the time label."""
        self.timeShow.setText(
            f"<html><head/><body><p><span style=\"font-size:16pt\">"
            f"{time // 60}[min] {time % 60:.02f}[s]"
            f"</span></p></body></html>"
        )

    def __add_result(self, result: Dict[str, Any]) -> None:
        """Add result items, except add to the list."""
        item = QListWidgetItem(result['Algorithm'])
        interrupt = result['interrupted']
        if interrupt == 'False':
            interrupt_icon = "task_completed.png"
        elif interrupt == 'N/A':
            interrupt_icon = "question.png"
        else:
            interrupt_icon = "interrupted.png"
        item.setIcon(QIcon(QPixmap(f":/icons/{interrupt_icon}")))
        if interrupt == 'False':
            interrupt_text = "No interrupt."
        else:
            interrupt_text = f"Interrupt at: {interrupt}"
        text = f"{result['Algorithm']} ({interrupt_text})"
        if interrupt == 'N/A':
            text += "\n※Completeness is unknown."
        item.setToolTip(text)
        self.result_list.addItem(item)

    @Slot(name='on_delete_button_clicked')
    def __delete_result(self) -> None:
        """Delete a result."""
        row = self.result_list.currentRow()
        if not row > -1:
            return

        if QMessageBox.question(
            self,
            "Delete",
            "Delete this result from list?"
        ) != QMessageBox.Yes:
            return

        self.mechanism_data.pop(row)
        self.result_list.takeItem(row)
        self.workbook_no_save()
        self.__has_result()

    @Slot(QModelIndex, name='on_result_list_clicked')
    def __has_result(self, *_) -> None:
        """Set enable if there has any result."""
        enable = self.result_list.currentRow() > -1
        for button in (
            self.merge_button,
            self.delete_button,
            self.result_load_settings,
            self.result_clipboard
        ):
            button.setEnabled(enable)

    @Slot(QModelIndex, name='on_result_list_doubleClicked')
    def __show_result(self, _=None) -> None:
        """Double click result item can show up preview dialog."""
        row = self.result_list.currentRow()
        if not row > -1:
            return

        dlg = PreviewDialog(self.mechanism_data[row], self.__get_path(row), self)
        dlg.show()
        dlg.exec_()
        dlg.deleteLater()

    @Slot(name='on_merge_button_clicked')
    def __merge_result(self) -> None:
        """Merge mechanism into main canvas."""
        row = self.result_list.currentRow()
        if not row > -1:
            return

        if QMessageBox.question(
            self,
            "Merge",
            "Add the result expression into storage?"
        ) == QMessageBox.Yes:
            expression: str = self.mechanism_data[row]['Expression']
            self.merge_result(expression, self.__get_path(row))

    def __get_path(self, row: int) -> List[List[_Coord]]:
        """Using result data to generate paths of mechanism."""
        result = self.mechanism_data[row]
        expression: str = result['Expression']
        same: Dict[int: int] = result['same']
        inputs: List[Tuple[int, int]] = result['input']
        input_list = []
        for b, d in inputs:
            for i in range(b):
                if i in same:
                    b -= 1
            for i in range(d):
                if i in same:
                    d -= 1
            input_list.append((b, d))

        vpoints = parse_vpoints(expression)
        expr = vpoints_configure(vpoints, input_list)
        b, d = input_list[0]
        base_angle = vpoints[b].slope_angle(vpoints[d])

        path = [[] for _ in range(len(vpoints))]
        for angle in range(360 + 1):
            try:
                result_list = expr_solving(
                    expr,
                    {i: f"P{i}" for i in range(len(vpoints))},
                    vpoints,
                    [base_angle + angle] + [0] * (len(input_list) - 1)
                )
            except ValueError:
                nan = float('nan')
                for i in range(len(vpoints)):
                    path[i].append((nan, nan))
            else:
                for i in range(len(vpoints)):
                    coord = result_list[i]
                    if type(coord[0]) is tuple:
                        path[i].append(coord[1])
                    else:
                        path[i].append(coord)
        return path

    @Slot(name='on_result_clipboard_clicked')
    def __copy_result_text(self) -> None:
        """Copy pretty print result as text."""
        QApplication.clipboard().setText(
            pprint.pformat(self.mechanism_data[self.result_list.currentRow()])
        )

    @Slot(name='on_save_profile_clicked')
    def __save_profile(self) -> None:
        """Save as new profile to collection widget."""
        if not self.mech_params:
            return
        name, ok = QInputDialog.getText(
            self,
            "Profile name",
            "Please enter the profile name:"
        )
        if not ok:
            return
        i = 0
        while (not name) and (name not in self.collections):
            name = f"Structure_{i}"
            i += 1

        mech_params = deepcopy(self.mech_params)
        for key in ('Placement', 'Target'):
            for mp in mech_params[key]:
                mech_params[key][mp] = None

        self.collections[name] = mech_params
        self.workbook_no_save()

    @Slot(name='on_load_profile_clicked')
    def __load_profile(self) -> None:
        """Load profile from collections dialog."""
        dlg = CollectionsDialog(
            self.collections,
            self.get_collection,
            self.workbook_no_save,
            self.prefer.tick_mark_option,
            self.preview_canvas.monochrome,
            self
        )
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return

        params = dlg.params
        name = dlg.name
        dlg.deleteLater()
        del dlg

        # Check the profile
        if not (params['Target'] and params['input'] and params['Placement']):
            QMessageBox.warning(
                self,
                "Invalid profile",
                "The profile is not configured yet."
            )
            return

        self.__set_profile(name, params)

    def __set_profile(self, profile_name: str, params: Dict[str, Any]) -> None:
        """Set profile to sub-widgets."""
        self.__clear_settings()
        self.mech_params = deepcopy(params)
        expression: str = self.mech_params['Expression']
        self.expression_string.setText(expression)
        target: Dict[int, List[_Coord]] = self.mech_params['Target']
        for name in sorted(target):
            self.target_points.addItem(f"P{name}")
            path = target[name]
            if path:
                self.path[name] = path.copy()
            else:
                self.path[name] = []
        if self.has_target():
            self.target_points.setCurrentRow(0)
        self.target_label.setVisible(self.has_target())

        # Parameter of link length and input angle.
        link_list = set()
        for vlink in parse_vlinks(expression):
            if len(vlink.points) < 2 or vlink.name == "ground":
                continue
            a = vlink.points[0]
            b = vlink.points[1]
            link_list.add(f"P{a}<->P{b}")
            for c in vlink.points[2:]:
                for d in (a, b):
                    link_list.add(f"P{c}<->P{d}")
        link_count = len(link_list)

        angle_list = set()
        input_list: List[Tuple[int, int]] = self.mech_params['input']
        for b, d in input_list:
            angle_list.add(f"P{b}->P{d}")
        angle_count = len(angle_list)

        self.parameter_list.setRowCount(0)
        placement: Dict[int, Tuple[float, float, float]] = self.mech_params['Placement']
        self.parameter_list.setRowCount(len(placement) + link_count + angle_count)

        def spinbox(
            v: float,
            *,
            minimum: float = 0.,
            maximum: float = 9999.,
            prefix: bool = False
        ) -> QDoubleSpinBox:
            double_spinbox = QDoubleSpinBox()
            double_spinbox.setMinimum(minimum)
            double_spinbox.setMaximum(maximum)
            double_spinbox.setSingleStep(10.0)
            double_spinbox.setValue(v)
            if prefix:
                double_spinbox.setPrefix("±")
            return double_spinbox

        # Position
        self.preview_canvas.from_profile(self.mech_params)
        pos_list = parse_pos(expression)
        for node, ref in sorted(self.preview_canvas.same.items()):
            pos_list.insert(node, pos_list[ref])

        row = 0
        for name in sorted(placement):
            coord = placement[name]
            self.parameter_list.setItem(row, 0, QTableWidgetItem(f"P{name}"))
            self.parameter_list.setItem(row, 1, QTableWidgetItem('Placement'))
            x, y = self.preview_canvas.pos[name]
            for i, s in enumerate([
                spinbox(coord[0] if coord else x, minimum=-9999.),
                spinbox(coord[1] if coord else y, minimum=-9999.),
                spinbox(coord[2] if coord else 25., prefix=True),
            ]):
                s.valueChanged.connect(self.update_range)
                self.parameter_list.setCellWidget(row, i + 2, s)
            row += 1

        # Default value of upper and lower
        for name in ('upper', 'lower'):
            if name not in self.mech_params:
                self.mech_params[name] = [0.] * (link_count + angle_count)

        upper_list: List[float] = self.mech_params['upper']
        lower_list: List[float] = self.mech_params['lower']

        def set_by_center(
            index: int,
            get_range: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by center."""

            @Slot(float)
            def func(value: float) -> None:
                range_value = get_range()
                upper_list[index] = value + range_value
                lower_list[index] = value - range_value

            return func

        def set_by_range(
            index: int,
            get_value: Callable[[], float]
        ) -> Callable[[float], None]:
            """Return a slot function use to set limit value by range."""

            @Slot(float)
            def func(value: float) -> None:
                center = get_value()
                upper_list[index] = center + value
                lower_list[index] = center - value

            return func

        for i, name in enumerate(sorted(link_list) + sorted(angle_list)):
            name_item = QTableWidgetItem(name)
            name_item.setToolTip(name)
            self.parameter_list.setItem(row, 0, name_item)
            upper = upper_list[i]
            lower = lower_list[i]
            if name in link_list:
                type_name = "Link"
                n1, n2 = (int(n.replace('P', '')) for n in name.split('<->'))
                link_length = self.preview_canvas.distance(n1, n2)
                if upper == 0.:
                    upper = link_length + 50
                if lower <= 0.:
                    lower = link_length - 50
                    lower = 0. if lower < 0 else lower
            else:
                type_name = "Input"
                if upper == 0.:
                    upper = 360.
            self.parameter_list.setItem(row, 1, QTableWidgetItem(type_name))
            # Set values
            upper_list[i] = upper
            lower_list[i] = lower
            # Spin box
            error_range = (upper - lower) / 2
            default_value = error_range + lower
            if name in link_list:
                s1 = spinbox(default_value)
            else:
                s1 = spinbox(default_value, maximum=360.)
            self.parameter_list.setCellWidget(row, 2, s1)
            s2 = spinbox(error_range, prefix=True)
            self.parameter_list.setCellWidget(row, 4, s2)
            # Signals
            s1.valueChanged.connect(set_by_center(i, s2.value))
            s2.valueChanged.connect(set_by_range(i, s1.value))
            row += 1

        self.update_range()

        self.profile_name.setText(profile_name)

        # Default value of algorithm option.
        if 'settings' in self.mech_params:
            self.alg_options.update(self.mech_params['settings'])
        else:
            self.__set_algorithm_default()

        self.__able_to_generate()

    @Slot(name='on_result_load_settings_clicked')
    def __load_result_settings(self) -> None:
        """Load settings from a result."""
        self.__has_result()
        row = self.result_list.currentRow()
        if not row > -1:
            return
        self.__clear_settings()
        result = self.mechanism_data[row]
        if result['Algorithm'] == str(AlgorithmType.RGA):
            self.type0.setChecked(True)
        elif result['Algorithm'] == str(AlgorithmType.Firefly):
            self.type1.setChecked(True)
        elif result['Algorithm'] == str(AlgorithmType.DE):
            self.type2.setChecked(True)
        # Copy to mechanism params.
        self.__set_profile("External setting", result)
        self.__set_time(result['time'])
        # Load settings.
        self.alg_options.clear()
        self.alg_options.update(result['settings'])

    @Slot(name='on_type0_clicked')
    @Slot(name='on_type1_clicked')
    @Slot(name='on_type2_clicked')
    def __set_algorithm_default(self) -> None:
        """Set the algorithm settings to default."""
        self.alg_options.clear()
        self.alg_options.update(defaultSettings)
        if self.type0.isChecked():
            self.alg_options.update(GeneticPrams)
        elif self.type1.isChecked():
            self.alg_options.update(FireflyPrams)
        elif self.type2.isChecked():
            self.alg_options.update(DifferentialPrams)

    @Slot(name='on_advance_button_clicked')
    def __show_advance(self) -> None:
        """Get the settings from advance dialog."""
        if self.type0.isChecked():
            type_num = AlgorithmType.RGA
        elif self.type1.isChecked():
            type_num = AlgorithmType.Firefly
        else:
            type_num = AlgorithmType.DE
        dlg = AlgorithmOptionDialog(type_num, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return

        self.alg_options['report'] = dlg.report.value()
        self.alg_options.pop('max_gen', None)
        self.alg_options.pop('min_fit', None)
        self.alg_options.pop('max_time', None)
        if dlg.max_gen_option.isChecked():
            self.alg_options['max_gen'] = dlg.max_gen.value()
        elif dlg.min_fit_option.isChecked():
            self.alg_options['min_fit'] = dlg.min_fit.value()
        elif dlg.max_time_option.isChecked():
            # Three spinbox value translate to second.
            self.alg_options['max_time'] = (
                dlg.max_time_h.value() * 3600
                + dlg.max_time_m.value() * 60
                + dlg.max_time_s.value()
            )
        else:
            raise ValueError("invalid option")

        def from_table(row: int) -> Union[int, float]:
            """Get algorithm data from table."""
            return dlg.alg_table.cellWidget(row, 1).value()

        pop_size = dlg.pop_size.value()
        if type_num == AlgorithmType.RGA:
            self.alg_options['nPop'] = pop_size
            for i, tag in enumerate(('pCross', 'pMute', 'pWin', 'bDelta')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.Firefly:
            self.alg_options['n'] = pop_size
            for i, tag in enumerate(('alpha', 'beta_min', 'gamma', 'beta0')):
                self.alg_options[tag] = from_table(i)
        elif type_num == AlgorithmType.DE:
            self.alg_options['NP'] = pop_size
            for i, tag in enumerate(('strategy', 'F', 'CR')):
                self.alg_options[tag] = from_table(i)

        dlg.deleteLater()

    @Slot()
    def update_range(self) -> None:
        """Update range values to main canvas."""

        def t(x: int, y: int) -> Union[str, float]:
            item = self.parameter_list.item(x, y)
            if item is None:
                return self.parameter_list.cellWidget(x, y).value()
            else:
                return item.text()

        self.update_ranges({
            t(row, 0): (t(row, 2), t(row, 3), t(row, 4))
            for row in range(self.parameter_list.rowCount())
            if t(row, 1) == 'Placement'
        })

    @Slot(name='on_expr_copy_clicked')
    def __copy_expr(self) -> None:
        """Copy profile expression."""
        text = self.expression_string.text()
        if text:
            QApplication.clipboard().setText(text)
