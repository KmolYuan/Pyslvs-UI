# -*- coding: utf-8 -*-

"""'dimensional_synthesis' module contains
dimensional synthesis functional interfaces.
"""

from __future__ import annotations

__all__ = ['Optimizer']
__author__ = "Yuan Chang"
__copyright__ = "Copyright (C) 2016-2021"
__license__ = "AGPL"
__email__ = "pyslvs@gmail.com"

from pprint import pformat
from copy import deepcopy
from math import hypot
from typing import (
    Any, Callable, Dict, Mapping, Iterable, Iterator, List, Optional,
    Sequence, TYPE_CHECKING, Tuple, Union, cast,
)
from lark.exceptions import LarkError
from openpyxl import load_workbook
from pyslvs import expr_solving, parse_pos, parse_vpoints, t_config
from pyslvs.optimization import norm_path
from pyslvs.metaheuristics import AlgorithmType, default
from qtpy.QtCore import QModelIndex, Slot
from qtpy.QtGui import QIcon, QPixmap
from qtpy.QtWidgets import (
    QApplication,
    QDoubleSpinBox,
    QHeaderView,
    QInputDialog,
    QListWidgetItem,
    QMessageBox,
    QRadioButton,
    QTableWidgetItem,
    QWidget,
)
from pyslvs_ui.graphics import PreviewCanvas, parse_path
from pyslvs_ui.synthesis import CollectionsDialog
from .dialogs import (
    AlgorithmOptionDialog,
    ChartDialog,
    EditPathDialog,
    PreviewDialog,
    ProgressDialog,
)
from .dimension_widget_ui import Ui_Form

if TYPE_CHECKING:
    from pyslvs_ui.widgets import MainWindowBase

_Pair = Tuple[int, int]
_Coord = Tuple[float, float]
_Range = Tuple[float, float, float]


class Optimizer(QWidget, Ui_Form):
    """Dimensional synthesis widget.

    User can run the dimensional synthesis here.
    """
    mech: Dict[str, Any]
    path: Dict[int, List[_Coord]]
    mechanism_data: List[Dict[str, Any]]
    alg_options: Dict[str, Union[int, float]]
    algorithm_options: Dict[AlgorithmType, QRadioButton]

    def __init__(self, parent: MainWindowBase):
        """Reference names:

        + Iteration collections.
        + Result data.
        + Main window function references.
        """
        super(Optimizer, self).__init__(parent)
        self.setupUi(self)
        self.mech = {}
        self.path = {}
        # Some reference of 'collections'
        self.collections = parent.collections.configure_widget.collections
        self.get_collection = parent.get_configure
        self.input_from = parent.input_from
        self.output_to = parent.output_to
        self.save_reply_box = parent.save_reply_box
        self.project_no_save = parent.project_no_save
        self.merge_result = parent.merge_result
        self.update_ranges = parent.main_canvas.update_ranges
        self.set_solving_path = parent.main_canvas.set_solving_path
        self.get_zoom = parent.main_canvas.get_zoom
        self.prefer = parent.prefer
        # Data and functions
        self.mechanism_data = []
        self.alg_options = default(AlgorithmType.DE)
        # Canvas
        self.preview_canvas = PreviewCanvas(self)
        self.preview_layout.addWidget(self.preview_canvas)
        # Splitter
        self.main_splitter.setStretchFactor(0, 10)
        self.main_splitter.setStretchFactor(1, 100)
        self.canvas_splitter.setSizes([80, 100])
        self.down_splitter.setSizes([20, 80])
        # Table widget column width
        header = self.parameter_list.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        self.algorithm_options = {}
        for opt in AlgorithmType:
            button = QRadioButton(opt.value, self)
            button.clicked.connect(self.__set_algorithm_default)
            self.algorithm_options[opt] = button
            self.algorithm_layout.addWidget(button)
        self.clear()

    def clear(self) -> None:
        """Clear all sub-widgets."""
        self.mechanism_data.clear()
        self.result_list.clear()
        self.__clear_settings()
        self.__has_result()

    def __clear_settings(self) -> None:
        """Clear sub-widgets that contain the setting."""
        self.clear_path(ask=False)
        self.path.clear()
        self.mech.clear()
        self.preview_canvas.clear()
        self.profile_name.clear()
        self.algorithm_options[AlgorithmType.DE].setChecked(True)
        self.__set_algorithm_default()
        self.parameter_list.setRowCount(0)
        self.target_points.clear()
        self.target_label.setVisible(self.has_target())
        self.expression_string.clear()
        self.update_range()
        self.__able_to_generate()

    def has_target(self) -> bool:
        """Return true if the panel is no target settings."""
        return self.target_points.count() > 0

    @Slot(name='on_clear_btn_clicked')
    def __user_clear(self) -> None:
        if not self.profile_name.text():
            return
        if QMessageBox.question(
            self,
            "Clear setting",
            "Do you want to clear the setting?"
        ) == QMessageBox.Yes:
            self.__clear_settings()

    def load_results(self, mechanism_data: Sequence[Mapping[str, Any]]) -> None:
        """Append results of project database to memory."""
        for e in mechanism_data:
            self.mechanism_data.append(dict(e))
            self.__add_result(e)

    def __current_path_changed(self) -> None:
        """Call the canvas to update to current target path."""
        self.set_solving_path({n: tuple(p) for n, p in self.path.items()})
        self.__able_to_generate()

    def current_path(self) -> List[_Coord]:
        """Return the pointer of current target path."""
        item = self.target_points.currentItem()
        if item is None:
            return []
        return self.path[int(item.text().replace('P', ''))]

    @Slot(str, name='on_target_points_currentTextChanged')
    def __set_target(self, _=None) -> None:
        """Switch to the current target path."""
        self.path_list.clear()
        for x, y in self.current_path():
            self.path_list.addItem(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @Slot(name='on_path_clear_clicked')
    def clear_path(self, *, ask: bool = True) -> None:
        """Clear the current target path."""
        if ask:
            if QMessageBox.question(
                self,
                "Clear path",
                "Are you sure to clear the current path?"
            ) != QMessageBox.Yes:
                return
        self.current_path().clear()
        self.path_list.clear()
        self.__current_path_changed()

    @Slot(name='on_path_copy_clicked')
    def __copy_path(self) -> None:
        """Copy the current path coordinates to clipboard."""
        if self.copy_as_csv.isChecked():
            text = '\n'.join(f"{x},{y}" for x, y in self.current_path())
        elif self.copy_as_array.isChecked():
            text = '\n'.join(f"[{x}, {y}]," for x, y in self.current_path())
        else:
            raise ValueError("invalid option")
        QApplication.clipboard().setText(text)

    @Slot(name='on_path_paste_clicked')
    def __paste_path(self) -> None:
        """Paste path data from clipboard."""
        self.__read_path_from_csv(QApplication.clipboard().text())

    @Slot(name='on_import_csv_btn_clicked')
    def __import_csv(self) -> None:
        """Paste path data from a text file."""
        file_name = self.input_from(
            "path data",
            ["Text file (*.txt)", "Comma-Separated Values (*.csv)"]
        )
        if not file_name:
            return
        with open(file_name, 'r', encoding='utf-8', newline='') as f:
            data = f.read()
        self.__read_path_from_csv(data)

    def __read_path_from_csv(self, raw: str, *, clear: bool = True) -> None:
        """Turn string to float then add them to current target path."""
        try:
            path = parse_path(raw)
        except LarkError as e:
            QMessageBox.warning(self, "Wrong format", f"{e}")
        else:
            self.set_path(path, clear=clear)

    @Slot(name='on_append_path_btn_clicked')
    def __append_path(self):
        """Append path from text."""
        raw, ok = QInputDialog.getMultiLineText(self, "Append path",
                                                "Path from csv format.")
        if ok and raw:
            self.__read_path_from_csv(raw, clear=False)

    @Slot(name='on_save_path_btn_clicked')
    def __save_path(self):
        """Save current path."""
        path = self.current_path()
        if not path:
            return
        file_name = self.output_to("Path file", ["Text file (*.txt)"])
        if not file_name:
            return
        with open(file_name, 'w+', encoding='utf-8') as f:
            f.write("\n".join(f"{x}, {y}" for x, y in path))
        self.save_reply_box("Path file", file_name)

    @Slot(name='on_import_xlsx_btn_clicked')
    def __import_xlsx(self) -> None:
        """Paste path data from a Excel file."""
        file_name = self.input_from(
            "Excel project",
            ["Microsoft Office Excel (*.xlsx *.xlsm *.xltx *.xltm)"]
        )
        if not file_name:
            return
        wb = load_workbook(file_name)
        sheets = wb.get_sheet_names()
        name, ok = QInputDialog.getItem(self, "Sheets", "Select a sheet:",
                                        sheets, 0)
        if not ok:
            return

        def get_path(sheet: str) -> Iterator[_Coord]:
            """Keep finding until there is no value"""
            ws = wb.get_sheet_by_name(sheets.index(sheet))
            i = 1
            while True:
                sx = ws.cell(i, 1).value
                sy = ws.cell(i, 2).value
                if None in {sx, sy}:
                    break
                try:
                    yield float(sx), float(sy)
                except Exception as e:
                    QMessageBox.warning(self, "File error", f"{e}")
                    return
                i += 1

        self.set_path(get_path(name))

    @Slot(name='on_edit_path_btn_clicked')
    def __adjust_path(self) -> None:
        """Show up path adjust dialog and
        get back the changes of current target path.
        """
        dlg = EditPathDialog(self)
        dlg.show()
        dlg.exec_()
        dlg.deleteLater()
        self.__current_path_changed()

    @Slot(name='on_norm_path_btn_clicked')
    def __norm_path(self) -> None:
        """Normalize current path."""
        scale, ok = QInputDialog.getDouble(
            self,
            "Scale",
            "Length of unit vector:",
            60, 0.01, 1000, 2)
        if ok:
            self.set_path(norm_path(self.current_path(), scale))

    def add_point(self, x: float, y: float) -> None:
        """Add path data to list widget and current target path."""
        self.current_path().append((x, y))
        self.path_list.addItem(f"({x:.08f}, {y:.08f})")
        self.path_list.setCurrentRow(self.path_list.count() - 1)
        self.__current_path_changed()

    def set_path(self, path: Iterable[_Coord], *, clear: bool = True) -> None:
        """Set the current path."""
        if clear:
            self.clear_path(ask=False)
        for x, y in path:
            self.add_point(x, y)
        self.__current_path_changed()

    @Slot(float, float)
    def set_point(self, x: float, y: float) -> None:
        """Set the coordinate of current target path."""
        if not self.edit_target_point_btn.isChecked():
            return
        for i, (cx, cy) in enumerate(self.current_path()):
            if hypot(x - cx, y - cy) < 10 / self.get_zoom():
                index = i
                self.path_list.setCurrentRow(index)
                break
        else:
            return
        self.current_path()[index] = (x, y)
        self.path_list.item(index).setText(f"({x:.04f}, {y:.04f})")
        self.__current_path_changed()

    @Slot(name='on_point_up_clicked')
    def __move_up_point(self) -> None:
        """Target point move up."""
        row = self.path_list.currentRow()
        if not (row > 0 and self.path_list.count() > 1):
            return
        path = self.current_path()
        path.insert(row - 1, (path[row][0], path[row][1]))
        path.pop(row + 1)
        c = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row - 1, f"({c[0]}, {c[1]})")
        self.path_list.takeItem(row + 1)
        self.path_list.setCurrentRow(row - 1)
        self.__current_path_changed()

    @Slot(name='on_point_down_clicked')
    def __move_down_point(self) -> None:
        """Target point move down."""
        row = self.path_list.currentRow()
        if not (
            row < self.path_list.count() - 1
            and self.path_list.count() > 1
        ):
            return
        path = self.current_path()
        path.insert(row + 2, (path[row][0], path[row][1]))
        path.pop(row)
        c = self.path_list.currentItem().text()[1:-1].split(", ")
        self.path_list.insertItem(row + 2, f"{c[0]}, {c[1]}")
        self.path_list.takeItem(row)
        self.path_list.setCurrentRow(row + 1)
        self.__current_path_changed()

    @Slot(name='on_point_delete_clicked')
    def __delete_point(self) -> None:
        """Delete a target point."""
        row = self.path_list.currentRow()
        if not row > -1:
            return
        self.current_path().pop(row)
        self.path_list.takeItem(row)
        self.__current_path_changed()

    def __able_to_generate(self) -> None:
        """Set button enable if all the data are already."""
        self.point_num.setText(
            "<p><span style=\"font-size:12pt;"
            f"color:#00aa00;\">{self.path_list.count()}</span></p>"
        )
        n = bool(
            self.mech
            and self.path_list.count() > 2
            and self.expression_string.text()
        )
        for button in (
            self.save_path_btn,
            self.edit_path_btn,
            self.norm_path_btn,
            self.synthesis_btn,
        ):
            button.setEnabled(n)

    @Slot(name='on_synthesis_btn_clicked')
    def __synthesis(self) -> None:
        """Start synthesis."""
        # Check if the amount of the target points are same
        length = -1
        for path in self.path.values():
            if length < 0:
                length = len(path)
            if len(path) != length:
                QMessageBox.warning(
                    self,
                    "Target Error",
                    "The length of target paths should be the same."
                )
                return
        # Get the algorithm type
        for option, button in self.algorithm_options.items():
            if button.isChecked():
                algorithm = option
                break
        else:
            raise ValueError("no option")
        mech = deepcopy(self.mech)
        mech['shape_only'] = self.shape_only_option.isChecked()
        if mech['shape_only']:
            if QMessageBox.question(
                self,
                "Elliptical Fourier Descriptor",
                "An even distribution will make the comparison more accurate.\n"
                "Do you make sure yet?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            ) == QMessageBox.No:
                return
        mech['expression'] = parse_vpoints(mech.pop('expression', []))
        mech['target'] = deepcopy(self.path)

        def name_in_table(target_name: str) -> int:
            """Find a target_name and return the row from the table."""
            for r in range(self.parameter_list.rowCount()):
                if self.parameter_list.item(r, 0).text() == target_name:
                    return r
            return -1

        placement: Dict[int, Tuple[float, float, float]] = mech['placement']
        for name in placement:
            row = name_in_table(f"P{name}")
            placement[name] = (
                self.parameter_list.cellWidget(row, 2).value(),
                self.parameter_list.cellWidget(row, 3).value(),
                self.parameter_list.cellWidget(row, 4).value(),
            )
        # Start progress dialog
        dlg = ProgressDialog(algorithm, mech, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return
        mechanisms_plot: List[Mapping[str, Any]] = []
        for data in dlg.mechanisms:
            mechanisms_plot.append({
                'time_fitness': data.pop('time_fitness'),
                'algorithm': data['algorithm'],
            })
            self.mechanism_data.append(data)
            self.__add_result(data)
        self.__set_time(dlg.time_spend)
        self.project_no_save()
        dlg.deleteLater()
        dlg = ChartDialog("Convergence Data", mechanisms_plot, self)
        dlg.show()
        dlg.exec_()
        dlg.deleteLater()

    def __set_time(self, time: float) -> None:
        """Set the time label."""
        self.timeShow.setText(
            f"<html><head/><body><p><span style=\"font-size:16pt\">"
            f"{int(time // 60):02d}min {time % 60:05.02f}s"
            f"</span></p></body></html>"
        )

    def __add_result(self, result: Mapping[str, Any]) -> None:
        """Add result items, except add to the list."""
        item = QListWidgetItem(result['algorithm'])
        interrupt = result['interrupted']
        if interrupt == 'False':
            interrupt_icon = "task_completed.png"
        elif interrupt == 'N/A':
            interrupt_icon = "question.png"
        else:
            interrupt_icon = "interrupted.png"
        item.setIcon(QIcon(QPixmap(f"icons:{interrupt_icon}")))
        if interrupt == 'False':
            interrupt_text = "No interrupt."
        else:
            interrupt_text = f"Interrupt at: {interrupt}"
        text = f"{result['algorithm']} ({interrupt_text})"
        if interrupt == 'N/A':
            text += "\n※Completeness is unknown."
        item.setToolTip(text)
        self.result_list.addItem(item)

    @Slot(name='on_delete_btn_clicked')
    def __delete_result(self) -> None:
        """Delete a result."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        if QMessageBox.question(
            self,
            "Delete",
            "Delete this result from list?"
        ) != QMessageBox.Yes:
            return
        self.mechanism_data.pop(row)
        self.result_list.takeItem(row)
        self.project_no_save()
        self.__has_result()

    @Slot(QModelIndex, name='on_result_list_clicked')
    def __has_result(self, *_) -> None:
        """Set enable if there has any result."""
        enable = self.result_list.currentRow() > -1
        for button in (
            self.merge_btn,
            self.delete_btn,
            self.result_load_settings,
            self.result_clipboard
        ):
            button.setEnabled(enable)

    @Slot(QModelIndex, name='on_result_list_doubleClicked')
    def __show_result(self, _=None) -> None:
        """Double click result item can show up preview dialog."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        dlg = PreviewDialog(self.mechanism_data[row], self.__get_path(row),
                            self.preview_canvas.monochrome, self)
        dlg.show()
        dlg.exec_()
        dlg.deleteLater()

    @Slot(name='on_merge_btn_clicked')
    def __merge_result(self) -> None:
        """Merge mechanism into main canvas."""
        row = self.result_list.currentRow()
        if not row > -1:
            return
        if QMessageBox.question(
            self,
            "Merge",
            "Add the result expression into storage?"
        ) == QMessageBox.Yes:
            expression: str = self.mechanism_data[row]['expression']
            self.merge_result(expression, self.__get_path(row))

    def __get_path(self, row: int) -> List[List[_Coord]]:
        """Using result data to generate paths of mechanism."""
        result = self.mechanism_data[row]
        expression: str = result['expression']
        same: Mapping[int, int] = result['same']
        inputs: List[Tuple[_Pair, _Coord]] = result['input']
        input_list = []
        for (b, d), _ in inputs:
            for i in range(b):
                if i in same:
                    b -= 1
            for i in range(d):
                if i in same:
                    d -= 1
            input_list.append((b, d))
        vpoints = parse_vpoints(expression)
        expr = t_config(vpoints, input_list)
        b, d = input_list[0]
        base_angle = vpoints[b].slope_angle(vpoints[d])
        path: List[List[_Coord]] = [[] for _ in range(len(vpoints))]
        input_pair = {(b, d): 0. for b, d in input_list}
        for angle in range(360 + 1):
            input_pair[b, d] = base_angle + angle
            try:
                result_list = expr_solving(expr, vpoints, input_pair)
            except ValueError:
                nan = float('nan')
                for i in range(len(vpoints)):
                    path[i].append((nan, nan))
            else:
                for i in range(len(vpoints)):
                    coord = result_list[i]
                    if isinstance(coord[0], tuple):
                        path[i].append(coord[1])
                    else:
                        path[i].append(coord)
        return path

    @Slot(name='on_result_clipboard_clicked')
    def __copy_result_text(self) -> None:
        """Copy pretty print result as text."""
        QApplication.clipboard().setText(
            pformat(self.mechanism_data[self.result_list.currentRow()])
        )

    @Slot(name='on_save_profile_clicked')
    def __save_profile(self) -> None:
        """Save as new profile to collection widget."""
        if not self.mech:
            return
        name, ok = QInputDialog.getText(
            self,
            "Profile name",
            "Please enter the profile name:"
        )
        if not ok:
            return
        i = 0
        while (not name) and (name not in self.collections):
            name = f"Structure_{i}"
            i += 1
        mech = deepcopy(self.mech)
        for key in ('placement', 'target'):
            for mp in mech[key]:
                mech[key][mp] = None
        self.collections[name] = mech
        self.project_no_save()

    @Slot(name='on_load_profile_clicked')
    def __load_profile(self) -> None:
        """Load profile from collections dialog."""
        dlg = CollectionsDialog(
            self.collections,
            self.get_collection,
            self.project_no_save,
            self.prefer.tick_mark_option,
            self.preview_canvas.monochrome,
            self
        )
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return
        params = dlg.params
        name = dlg.name
        dlg.deleteLater()
        del dlg
        # Check the profile
        if not (params['target'] and params['input'] and params['placement']):
            QMessageBox.warning(
                self,
                "Invalid profile",
                "The profile is not configured yet."
            )
            return
        self.__set_profile(name, params)

    def __set_profile(self, profile_name: str,
                      params: Mapping[str, Any]) -> None:
        """Set profile to sub-widgets."""
        self.__clear_settings()
        self.mech = dict(deepcopy(params))
        expression: str = self.mech['expression']
        self.expression_string.setText(expression)
        target: Mapping[int, List[_Coord]] = self.mech['target']
        for p in sorted(target):
            self.target_points.addItem(f"P{p}")
            if target[p]:
                self.path[p] = target[p].copy()
            else:
                self.path[p] = []
        if self.has_target():
            self.target_points.setCurrentRow(0)
        self.target_label.setVisible(self.has_target())
        inputs: List[Tuple[_Pair, List[float]]] = self.mech.get('input', [])
        self.mech['input'] = inputs
        placement: Mapping[int, Optional[_Range]] = self.mech.get('placement',
                                                                  {})
        self.mech['placement'] = placement
        # Table settings
        self.parameter_list.setRowCount(0)
        self.parameter_list.setRowCount(len(inputs) + len(placement) + 1)
        row = 0

        def spinbox(
            v: float = 0.,
            *,
            minimum: float = 0.,
            maximum: float = 9999.,
            prefix: bool = False
        ) -> QDoubleSpinBox:
            double_spinbox = QDoubleSpinBox()
            double_spinbox.setMinimum(minimum)
            double_spinbox.setMaximum(maximum)
            double_spinbox.setValue(v)
            if prefix:
                double_spinbox.setPrefix("±")
            return double_spinbox

        def set_angle(index1: int, index2: int) -> Callable[[float], None]:
            """Return a slot function use to set angle value."""

            @Slot(float)
            def func(value: float) -> None:
                inputs[index1][1][index2] = value

            return func

        # Angles
        for i, ((b, d), se) in enumerate(inputs):
            self.parameter_list.setItem(row, 0, QTableWidgetItem(f"P{b}->P{d}"))
            self.parameter_list.setItem(row, 1, QTableWidgetItem('input'))
            s1 = spinbox(se[0], maximum=360.)
            s2 = spinbox(se[1], maximum=360.)
            self.parameter_list.setCellWidget(row, 2, s1)
            self.parameter_list.setCellWidget(row, 3, s2)
            s1.valueChanged.connect(set_angle(i, 0))
            s2.valueChanged.connect(set_angle(i, 1))
            row += 1
        # Grounded joints
        self.preview_canvas.from_profile(self.mech)
        pos_list = parse_pos(expression)
        for node, ref in sorted(self.preview_canvas.same.items()):
            pos_list.insert(node, pos_list[ref])
        for p in sorted(placement):
            coord = placement[p]
            self.parameter_list.setItem(row, 0, QTableWidgetItem(f"P{p}"))
            self.parameter_list.setItem(row, 1, QTableWidgetItem('placement'))
            x, y = self.preview_canvas.pos[p]
            for i, s in enumerate([
                spinbox(coord[0] if coord else x, minimum=-9999.),
                spinbox(coord[1] if coord else y, minimum=-9999.),
                spinbox(coord[2] if coord else 5., prefix=True),
            ]):
                s.valueChanged.connect(self.update_range)
                self.parameter_list.setCellWidget(row, i + 2, s)
            row += 1
        # Default value of upper and lower
        self.mech['upper'] = self.mech.get('upper', 100)
        self.mech['lower'] = self.mech.get('lower', 0)

        def set_link(opt: str) -> Callable[[float], None]:
            """Set link length."""

            @Slot(float)
            def func(value: float) -> None:
                self.mech[opt] = value

            return func

        self.parameter_list.setItem(row, 0, QTableWidgetItem("L"))
        self.parameter_list.setItem(row, 1, QTableWidgetItem('link'))
        for i, (s, tag) in enumerate([
            (spinbox(), 'upper'),
            (spinbox(), 'lower'),
        ]):
            s.setValue(self.mech[tag])
            s.valueChanged.connect(set_link(tag))
            self.parameter_list.setCellWidget(row, i + 2, s)
        # Update previews
        self.update_range()
        self.profile_name.setText(profile_name)
        # Default value of algorithm option
        if 'settings' in self.mech:
            self.alg_options.update(self.mech['settings'])
        else:
            self.__set_algorithm_default()
        self.__able_to_generate()

    @Slot(name='on_result_load_settings_clicked')
    def __load_result_settings(self) -> None:
        """Load settings from a result."""
        self.__has_result()
        row = self.result_list.currentRow()
        if not row > -1:
            return
        self.__clear_settings()
        result = self.mechanism_data[row]
        for option, button in self.algorithm_options.items():
            if result.get('algorithm', "") == option.value:
                button.setChecked(True)
                break
        else:
            raise ValueError("no option")
        # Copy to mechanism params
        self.__set_profile("External setting", result)
        self.__set_time(result.get('time', 0))
        # Load settings
        self.alg_options.clear()
        self.alg_options.update(result.get('settings', {}))
        self.shape_only_option.setChecked(result.get('shape_only', False))

    @Slot()
    def __set_algorithm_default(self) -> None:
        """Set the algorithm settings to default."""
        self.alg_options.clear()
        for opt, button in self.algorithm_options.items():
            if button.isChecked():
                self.alg_options.update(default(opt))

    @Slot(name='on_advance_btn_clicked')
    def __show_advance(self) -> None:
        """Get the settings from advance dialog."""
        for option, button in self.algorithm_options.items():
            if button.isChecked():
                algorithm = option
                break
        else:
            raise ValueError("no option")
        dlg = AlgorithmOptionDialog(algorithm, self.alg_options, self)
        dlg.show()
        if not dlg.exec_():
            dlg.deleteLater()
            return
        self.alg_options['report'] = dlg.report.value()
        self.alg_options.pop('max_gen', None)
        self.alg_options.pop('min_fit', None)
        self.alg_options.pop('max_time', None)
        if dlg.max_gen_option.isChecked():
            self.alg_options['max_gen'] = dlg.max_gen.value()
        elif dlg.min_fit_option.isChecked():
            self.alg_options['min_fit'] = dlg.min_fit.value()
        elif dlg.max_time_option.isChecked():
            # Three spinbox value translate to second
            self.alg_options['max_time'] = (
                dlg.max_time_h.value() * 3600
                + dlg.max_time_m.value() * 60
                + dlg.max_time_s.value()
            )
        else:
            raise ValueError("invalid option")
        self.alg_options['pop_num'] = dlg.pop_size.value()
        for row in range(dlg.alg_table.rowCount()):
            option = dlg.alg_table.item(row, 0).text()
            self.alg_options[option] = dlg.alg_table.cellWidget(row, 1).value()
        dlg.deleteLater()

    @Slot()
    def update_range(self) -> None:
        """Update range values to main canvas."""

        def t(row: int, col: int) -> Union[str, float]:
            item = self.parameter_list.item(row, col)
            if item is None:
                w: QDoubleSpinBox = self.parameter_list.cellWidget(row, col)
                return w.value()
            else:
                return item.text()

        self.update_ranges({
            cast(str, t(row, 0)): (
                cast(float, t(row, 2)),
                cast(float, t(row, 3)),
                cast(float, t(row, 4)),
            ) for row in range(self.parameter_list.rowCount())
            if t(row, 1) == 'placement'
        })

    @Slot(name='on_expr_copy_clicked')
    def __copy_expr(self) -> None:
        """Copy profile expression."""
        text = self.expression_string.text()
        if text:
            QApplication.clipboard().setText(text)
